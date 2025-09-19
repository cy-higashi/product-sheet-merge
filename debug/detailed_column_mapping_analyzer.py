#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
詳細カラムマッピング分析スクリプト

Phase3→Phase4→Phase5の各段階でのデータ変換を追跡し、
カラムズレの原因を特定する
"""

import os
import pandas as pd
from openpyxl import load_workbook
import json
from datetime import datetime
from pathlib import Path

class ColumnMappingAnalyzer:
    def __init__(self, municipality_name="2025-09-09_福岡県北九州市"):
        self.municipality = municipality_name
        self.base_dirs = {
            'phase1': f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase1\\HARV\\{municipality_name}",
            'phase2': f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase2\\HARV\\{municipality_name}",
            'phase3': f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase3\\HARV\\{municipality_name}"
        }
        self.results = {}
    
    def analyze_phase3_structure(self):
        """Phase3の転置データ構造を分析"""
        print("\n=== Phase3構造分析 ===")
        phase3_dir = self.base_dirs['phase3']
        
        if not os.path.exists(phase3_dir):
            print(f"Phase3ディレクトリが存在しません: {phase3_dir}")
            return
        
        pat_files = [f for f in os.listdir(phase3_dir) if f.startswith("PAT") and f.endswith(".xlsx")]
        print(f"PATファイル数: {len(pat_files)}")
        
        phase3_analysis = {}
        
        for pat_file in pat_files:
            file_path = os.path.join(phase3_dir, pat_file)
            print(f"\n--- {pat_file} ---")
            
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                
                # 構造情報
                structure_info = {
                    'max_row': ws.max_row,
                    'max_col': ws.max_column,
                    'sample_data': []
                }
                
                # 最初の10行のサンプルデータ
                for row in range(1, min(11, ws.max_row + 1)):
                    row_data = []
                    for col in range(1, min(11, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(cell_value)
                    structure_info['sample_data'].append(row_data)
                
                # ヘッダー候補の特定
                header_candidates = self.find_header_structure(ws)
                structure_info['header_candidates'] = header_candidates
                
                phase3_analysis[pat_file] = structure_info
                
                print(f"  サイズ: {ws.max_row}行 × {ws.max_column}列")
                print(f"  ヘッダー候補: {len(header_candidates)}個")
                
            except Exception as e:
                print(f"  エラー: {e}")
                phase3_analysis[pat_file] = {'error': str(e)}
        
        self.results['phase3_analysis'] = phase3_analysis
        return phase3_analysis
    
    def analyze_phase4_normalization(self):
        """Phase4の正規化処理を分析"""
        print("\n=== Phase4正規化分析 ===")
        phase3_dir = self.base_dirs['phase3']
        
        if not os.path.exists(phase3_dir):
            return
        
        normalized_files = [f for f in os.listdir(phase3_dir) if f.endswith("_normalized.xlsx")]
        print(f"正規化ファイル数: {len(normalized_files)}")
        
        phase4_analysis = {}
        
        for norm_file in normalized_files:
            file_path = os.path.join(phase3_dir, norm_file)
            print(f"\n--- {norm_file} ---")
            
            try:
                df = pd.read_excel(file_path)
                
                analysis_info = {
                    'shape': df.shape,
                    'columns': list(df.columns),
                    'sample_data': df.head(5).to_dict('records'),
                    'column_data_types': {col: str(df[col].dtype) for col in df.columns},
                    'null_counts': df.isnull().sum().to_dict()
                }
                
                # 特定の列での異常データをチェック
                suspicious_data = self.detect_suspicious_data(df)
                analysis_info['suspicious_data'] = suspicious_data
                
                phase4_analysis[norm_file] = analysis_info
                
                print(f"  形状: {df.shape}")
                print(f"  列数: {len(df.columns)}")
                if suspicious_data:
                    print(f"  疑わしいデータ: {len(suspicious_data)}件")
                
            except Exception as e:
                print(f"  エラー: {e}")
                phase4_analysis[norm_file] = {'error': str(e)}
        
        self.results['phase4_analysis'] = phase4_analysis
        return phase4_analysis
    
    def analyze_phase5_integration(self):
        """Phase5の統合処理を分析"""
        print("\n=== Phase5統合分析 ===")
        phase3_dir = self.base_dirs['phase3']
        all_collect_file = os.path.join(phase3_dir, "all_collect.xlsx")
        
        if not os.path.exists(all_collect_file):
            print(f"all_collect.xlsxが存在しません: {all_collect_file}")
            return
        
        try:
            df = pd.read_excel(all_collect_file)
            
            phase5_analysis = {
                'shape': df.shape,
                'columns': list(df.columns),
                'sample_data': df.head(10).to_dict('records'),
                'column_analysis': {}
            }
            
            # 各列の詳細分析
            for col in df.columns:
                col_analysis = {
                    'data_type': str(df[col].dtype),
                    'null_count': int(df[col].isnull().sum()),
                    'unique_count': int(df[col].nunique()),
                    'sample_values': df[col].dropna().head(10).tolist()
                }
                
                # 異常パターンの検出
                if '商品名' in col:
                    # 商品名列に人名が含まれているかチェック
                    name_patterns = ['河瀨', '担当者', '氏名']
                    suspicious_values = []
                    for pattern in name_patterns:
                        matches = df[df[col].astype(str).str.contains(pattern, na=False)]
                        if not matches.empty:
                            suspicious_values.extend(matches[col].tolist())
                    col_analysis['suspicious_values'] = suspicious_values
                
                elif '内容量' in col or '容量' in col:
                    # 内容量列に会社名が含まれているかチェック
                    company_patterns = ['株式会社', '有限会社', '合同会社', '企業']
                    suspicious_values = []
                    for pattern in company_patterns:
                        matches = df[df[col].astype(str).str.contains(pattern, na=False)]
                        if not matches.empty:
                            suspicious_values.extend(matches[col].tolist())
                    col_analysis['suspicious_values'] = suspicious_values
                
                phase5_analysis['column_analysis'][col] = col_analysis
            
            # ファイル別の分析
            if 'ファイル名' in df.columns:
                hamada_data = df[df['ファイル名'].astype(str).str.contains('ハマダ', na=False)]
                phase5_analysis['hamada_specific'] = {
                    'record_count': len(hamada_data),
                    'sample_records': hamada_data.head(5).to_dict('records')
                }
            
            self.results['phase5_analysis'] = phase5_analysis
            
            print(f"統合データ形状: {df.shape}")
            print(f"列数: {len(df.columns)}")
            print(f"ハマダ関連レコード: {len(hamada_data) if 'ファイル名' in df.columns else 'N/A'}")
            
            return phase5_analysis
            
        except Exception as e:
            print(f"Phase5分析エラー: {e}")
            return {'error': str(e)}
    
    def find_header_structure(self, ws):
        """ワークシートのヘッダー構造を特定"""
        header_candidates = []
        
        # 最初の20行を検索
        for row in range(1, min(21, ws.max_row + 1)):
            for col in range(1, min(21, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    # ヘッダーキーワード
                    if any(keyword in cell_value for keyword in 
                          ["返礼品コード", "商品名", "項目", "No.", "担当者", "会社名", "内容量"]):
                        header_candidates.append({
                            'row': row,
                            'col': col,
                            'value': cell_value,
                            'coordinate': ws.cell(row=row, column=col).coordinate
                        })
        
        return header_candidates
    
    def detect_suspicious_data(self, df):
        """疑わしいデータパターンを検出"""
        suspicious_data = []
        
        for col in df.columns:
            if '商品名' in col:
                # 商品名列に人名パターン
                name_patterns = ['河瀨', '担当者', '氏名', '様', 'さん']
                for pattern in name_patterns:
                    matches = df[df[col].astype(str).str.contains(pattern, na=False)]
                    for idx, row in matches.iterrows():
                        suspicious_data.append({
                            'type': 'name_in_product',
                            'column': col,
                            'row': int(idx),
                            'value': row[col],
                            'pattern': pattern
                        })
            
            elif '内容量' in col or '容量' in col:
                # 内容量列に会社名パターン
                company_patterns = ['株式会社', '有限会社', '合同会社']
                for pattern in company_patterns:
                    matches = df[df[col].astype(str).str.contains(pattern, na=False)]
                    for idx, row in matches.iterrows():
                        suspicious_data.append({
                            'type': 'company_in_volume',
                            'column': col,
                            'row': int(idx),
                            'value': row[col],
                            'pattern': pattern
                        })
        
        return suspicious_data
    
    def trace_specific_file_processing(self, target_filename):
        """特定ファイルの処理過程を追跡"""
        print(f"\n=== ファイル処理追跡: {target_filename} ===")
        
        # Phase1でのパターン検出
        phase1_file_pattern = os.path.join(self.base_dirs['phase1'], f"{self.municipality}_ファイル別パターン.xlsx")
        
        if os.path.exists(phase1_file_pattern):
            try:
                df1 = pd.read_excel(phase1_file_pattern)
                target_records = df1[df1['ファイル名'] == target_filename]
                
                if not target_records.empty:
                    pattern_name = target_records.iloc[0]['パターン名']
                    print(f"検出パターン: {pattern_name}")
                    
                    # Phase2でのパターン定義
                    phase2_file = os.path.join(self.base_dirs['phase2'], f"{self.municipality}_パターン一覧_Phase2.xlsx")
                    if os.path.exists(phase2_file):
                        df2 = pd.read_excel(phase2_file, header=None)
                        pattern_def = df2[df2[0] == pattern_name]
                        if not pattern_def.empty:
                            print(f"パターン定義: {list(pattern_def.iloc[0])}")
                    
                    # Phase3での転置結果
                    phase3_file = os.path.join(self.base_dirs['phase3'], f"{pattern_name}.xlsx")
                    if os.path.exists(phase3_file):
                        wb3 = load_workbook(phase3_file, data_only=True)
                        ws3 = wb3.active
                        print(f"Phase3転置結果: {ws3.max_row}行 × {ws3.max_column}列")
                        
                        # 該当ファイルのデータを検索
                        file_rows = []
                        for row in range(1, min(ws3.max_row + 1, 100)):
                            cell_value = ws3.cell(row=row, column=1).value
                            if cell_value and target_filename in str(cell_value):
                                row_data = []
                                for col in range(1, min(ws3.max_column + 1, 20)):
                                    row_data.append(ws3.cell(row=row, column=col).value)
                                file_rows.append(row_data)
                        
                        print(f"該当ファイルのデータ行: {len(file_rows)}行")
                        if file_rows:
                            print("サンプルデータ:")
                            for i, row in enumerate(file_rows[:3]):
                                print(f"  行{i+1}: {row[:10]}")  # 最初の10列
                
            except Exception as e:
                print(f"ファイル追跡エラー: {e}")
    
    def generate_recommendations(self):
        """カラムズレ問題の修正提案を生成"""
        print("\n=== 修正提案 ===")
        
        recommendations = []
        
        # Phase3分析結果から
        if 'phase3_analysis' in self.results:
            recommendations.append({
                'phase': 'Phase3',
                'issue': 'ヘッダー検出の精度向上',
                'recommendation': 'find_header_base_position関数で、複数の候補から最適な位置を選択する機能を追加'
            })
        
        # Phase4分析結果から
        if 'phase4_analysis' in self.results:
            for file, analysis in self.results['phase4_analysis'].items():
                if 'suspicious_data' in analysis and analysis['suspicious_data']:
                    recommendations.append({
                        'phase': 'Phase4',
                        'file': file,
                        'issue': f"疑わしいデータ {len(analysis['suspicious_data'])}件検出",
                        'recommendation': 'データクリーニングルールの追加'
                    })
        
        # Phase5分析結果から
        if 'phase5_analysis' in self.results:
            col_analysis = self.results['phase5_analysis'].get('column_analysis', {})
            for col, info in col_analysis.items():
                if 'suspicious_values' in info and info['suspicious_values']:
                    recommendations.append({
                        'phase': 'Phase5',
                        'column': col,
                        'issue': f"異常データ {len(info['suspicious_values'])}件",
                        'recommendation': f'{col}列の値検証とクリーニング処理を追加'
                    })
        
        self.results['recommendations'] = recommendations
        
        print("修正提案:")
        for i, rec in enumerate(recommendations, 1):
            print(f"{i}. [{rec['phase']}] {rec['issue']}")
            print(f"   → {rec['recommendation']}")
    
    def run_full_analysis(self):
        """完全分析の実行"""
        print("=== 詳細カラムマッピング分析開始 ===")
        print(f"対象自治体: {self.municipality}")
        print(f"分析開始時刻: {datetime.now()}")
        
        # 各フェーズの分析
        self.analyze_phase3_structure()
        self.analyze_phase4_normalization()
        self.analyze_phase5_integration()
        
        # 特定ファイルの追跡
        target_file = "2023-08-03_返礼品登録シート_株式会社ハマダ(赤身スライス900g定期便)（市確認）.xlsx"
        self.trace_specific_file_processing(target_file)
        
        # 修正提案の生成
        self.generate_recommendations()
        
        # 結果の保存
        output_file = "debug/detailed_column_mapping_analysis.json"
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(self.results, f, ensure_ascii=False, indent=2, default=str)
            print(f"\n分析結果を保存しました: {output_file}")
        except Exception as e:
            print(f"結果保存エラー: {e}")
        
        print("\n=== 分析完了 ===")
        return self.results

def main():
    analyzer = ColumnMappingAnalyzer()
    results = analyzer.run_full_analysis()
    return results

if __name__ == "__main__":
    main()
