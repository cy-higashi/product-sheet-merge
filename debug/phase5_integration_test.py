import os
from openpyxl import load_workbook, Workbook

# ===== 正常版（merge.py相当）の統合ロジック =====
def update_master_headers_simple(master_headers, source_headers):
    """
    シンプルな完全一致判定
    """
    for header in source_headers:
        if header not in master_headers:
            master_headers.append(header)
    return master_headers

def process_file_simple(file_path, master_headers):
    """
    正常版（merge.py相当）のファイル処理ロジック
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # ワークシート全体の値を2次元リストにコピー
    data = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        data.append([cell.value for cell in row])

    # ヘッダー行（B列が「項目」）を検索
    header_row_index = None
    for i, row in enumerate(data):
        if len(row) >= 2 and row[1] == "項目":
            header_row_index = i
            break

    if header_row_index is None:
        print(f"ファイル内に「項目」行が見つかりません: {os.path.basename(file_path)}")
        return [], master_headers

    # シンプルなヘッダー抽出（None値を除外）
    source_headers = [cell for cell in data[header_row_index][2:] if cell is not None]
    master_headers = update_master_headers_simple(master_headers, source_headers)

    # データ行処理
    output_rows = []
    for row in data[header_row_index+1:]:
        new_row = []
        # A列: ファイル名
        existing_file_name = row[0] if len(row) > 0 and row[0] is not None else os.path.basename(file_path)
        new_row.append(existing_file_name)
        # B列: 項目値
        value_B = row[1] if len(row) > 1 else None
        new_row.append(value_B)
        # C列以降: シンプルな照合
        for header in master_headers:
            if header in source_headers:
                idx = source_headers.index(header)
                new_row.append(row[2 + idx] if 2 + idx < len(row) else None)
            else:
                new_row.append(None)
        output_rows.append(new_row)

    return output_rows, master_headers

# ===== 問題版（error1_merge.py相当）の統合ロジック =====
def normalize_header(header):
    """
    ヘッダーを正規化します（改行→空白、連続空白圧縮、前後trim）
    """
    if header is None:
        return ""
    header_str = str(header)
    # 改行を空白に変換
    header_str = header_str.replace('\n', ' ').replace('\r', ' ')
    # 連続空白を単一空白に圧縮
    import re
    header_str = re.sub(r'\s+', ' ', header_str)
    # 前後trim
    header_str = header_str.strip()
    return header_str

def is_generic_group_header(header):
    """
    汎用グループ名かどうかを判定します
    """
    if not header:
        return True
    generic_patterns = [
        '返礼品発送元情報', '発送', '画像', '参考URL', '返礼品', '商品', '情報'
    ]
    normalized_header = normalize_header(header)
    return any(pattern in normalized_header for pattern in generic_patterns)

def create_effective_headers(main_headers, sub_headers):
    """
    メインヘッダーとサブヘッダーから実効ヘッダーを生成します
    """
    effective_headers = []
    max_len = max(len(main_headers), len(sub_headers)) if sub_headers else len(main_headers)
    
    for i in range(max_len):
        main_header = main_headers[i] if i < len(main_headers) else None
        sub_header = sub_headers[i] if i < len(sub_headers) else None
        
        # 配列位置を維持しながら実効ヘッダーを生成
        if main_header is not None and str(main_header).strip():
            main_normalized = normalize_header(main_header)
            if sub_header is not None and str(sub_header).strip() and not is_generic_group_header(main_header):
                # メインとサブの両方がある場合、合成キーを作成
                effective_header = f"{main_normalized}:{normalize_header(sub_header)}"
            else:
                # メインのみの場合
                effective_header = main_normalized
        elif sub_header is not None and str(sub_header).strip():
            # サブ見出しのみの場合
            effective_header = normalize_header(sub_header)
        else:
            # 両方とも空の場合は列番号を使用
            effective_header = f"列{i+3}"  # C列から開始なので+3
            
        effective_headers.append(effective_header)
    
    return effective_headers

def update_master_headers_complex(master_headers, source_headers):
    """
    複雑な正規化ヘッダー判定
    """
    for header in source_headers:
        if header not in master_headers:
            master_headers.append(header)
    return master_headers

def process_file_complex(file_path, master_headers):
    """
    問題版（error1_merge.py相当）のファイル処理ロジック
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # ワークシート全体の値を2次元リストにコピー
    data = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        data.append([cell.value for cell in row])

    # ヘッダー行（B列が「項目」）を検索
    header_row_index = None
    for i, row in enumerate(data):
        if len(row) >= 2 and row[1] == "項目":
            header_row_index = i
            break

    if header_row_index is None:
        print(f"ファイル内に「項目」行が見つかりません: {os.path.basename(file_path)}")
        return [], master_headers

    # 複雑なヘッダー抽出（メイン+サブ見出し）
    main_headers = data[header_row_index][2:]  # None値も保持
    sub_headers = []
    if header_row_index + 1 < len(data):
        sub_headers = data[header_row_index + 1][2:]  # None値も保持
    
    # 実効ヘッダーを生成
    source_headers = create_effective_headers(main_headers, sub_headers)
    master_headers = update_master_headers_complex(master_headers, source_headers)

    # データ行処理
    output_rows = []
    for row in data[header_row_index+1:]:
        new_row = []
        # A列: ファイル名
        existing_file_name = row[0] if len(row) > 0 and row[0] is not None else os.path.basename(file_path)
        new_row.append(existing_file_name)
        # B列: 項目値
        value_B = row[1] if len(row) > 1 else None
        new_row.append(value_B)
        # C列以降: 正規化された照合
        for header in master_headers:
            normalized_header = normalize_header(header)
            matching_idx = None
            for i, source_header in enumerate(source_headers):
                if normalize_header(source_header) == normalized_header:
                    matching_idx = i
                    break
            
            if matching_idx is not None:
                new_row.append(row[2 + matching_idx] if 2 + matching_idx < len(row) else None)
            else:
                new_row.append(None)
        output_rows.append(new_row)

    return output_rows, master_headers

# ===== 検証メイン関数 =====
def test_integration(municipality_name="不整合テスト自治体v3"):
    """
    両方の統合ロジックを比較検証
    """
    base_dir = os.path.join(
        r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase3\HARV',
        municipality_name
    )
    
    if not os.path.exists(base_dir):
        print(f"指定フォルダが存在しません: {base_dir}")
        return
    
    # 対象ファイルを取得
    files = [f for f in os.listdir(base_dir) if f.startswith("PAT") and f.endswith("_normalized.xlsx")]
    if not files:
        print("_normalized.xlsx ファイルが見つかりません。")
        return
    
    print(f"検証対象ファイル: {len(files)}個")
    print("=" * 80)
    
    # 正常版の処理
    print("=== 正常版（merge.py相当）の処理 ===")
    master_headers_simple = []
    master_data_rows_simple = []
    
    for file in files:
        file_path = os.path.join(base_dir, file)
        try:
            rows, master_headers_simple = process_file_simple(file_path, master_headers_simple)
            master_data_rows_simple.extend(rows)
            print(f"正常版処理完了: {file} (行数: {len(rows)})")
        except Exception as e:
            print(f"正常版エラー: {file} → {e}")
    
    # 正常版の出力
    simple_wb = Workbook()
    simple_ws = simple_wb.active
    simple_ws.title = "simple_version"
    
    # ヘッダー設定
    simple_ws.cell(row=1, column=1, value="ファイル名")
    simple_ws.cell(row=1, column=2, value="項目")
    for i, header in enumerate(master_headers_simple, start=3):
        simple_ws.cell(row=1, column=i, value=header)
    
    # データ設定
    current_row = 2
    for row in master_data_rows_simple:
        for j, value in enumerate(row, start=1):
            simple_ws.cell(row=current_row, column=j, value=value)
        current_row += 1
    
    simple_output_path = os.path.join(base_dir, "test_simple_integration.xlsx")
    simple_wb.save(simple_output_path)
    print(f"正常版出力完了: {simple_output_path}")
    print(f"正常版ヘッダー数: {len(master_headers_simple)}")
    print(f"正常版データ行数: {len(master_data_rows_simple)}")
    
    print("\n" + "=" * 80)
    
    # 問題版の処理
    print("=== 問題版（error1_merge.py相当）の処理 ===")
    master_headers_complex = []
    master_data_rows_complex = []
    
    for file in files:
        file_path = os.path.join(base_dir, file)
        try:
            rows, master_headers_complex = process_file_complex(file_path, master_headers_complex)
            master_data_rows_complex.extend(rows)
            print(f"問題版処理完了: {file} (行数: {len(rows)})")
        except Exception as e:
            print(f"問題版エラー: {file} → {e}")
    
    # 問題版の出力
    complex_wb = Workbook()
    complex_ws = complex_wb.active
    complex_ws.title = "complex_version"
    
    # ヘッダー設定
    complex_ws.cell(row=1, column=1, value="ファイル名")
    complex_ws.cell(row=1, column=2, value="項目")
    for i, header in enumerate(master_headers_complex, start=3):
        complex_ws.cell(row=1, column=i, value=header)
    
    # データ設定
    current_row = 2
    for row in master_data_rows_complex:
        for j, value in enumerate(row, start=1):
            complex_ws.cell(row=current_row, column=j, value=value)
        current_row += 1
    
    complex_output_path = os.path.join(base_dir, "test_complex_integration.xlsx")
    complex_wb.save(complex_output_path)
    print(f"問題版出力完了: {complex_output_path}")
    print(f"問題版ヘッダー数: {len(master_headers_complex)}")
    print(f"問題版データ行数: {len(master_data_rows_complex)}")
    
    print("\n" + "=" * 80)
    
    # 比較結果
    print("=== 比較結果 ===")
    print(f"ヘッダー数の差: {len(master_headers_complex) - len(master_headers_simple)}")
    print(f"データ行数の差: {len(master_data_rows_complex) - len(master_data_rows_simple)}")
    
    # ヘッダーの違いを詳細出力
    print("\n=== ヘッダーの違い ===")
    print("正常版のみにあるヘッダー:")
    for header in master_headers_simple:
        if header not in master_headers_complex:
            print(f"  - {header}")
    
    print("問題版のみにあるヘッダー:")
    for header in master_headers_complex:
        if header not in master_headers_simple:
            print(f"  + {header}")
    
    print("\n検証完了！")

if __name__ == "__main__":
    test_integration()
