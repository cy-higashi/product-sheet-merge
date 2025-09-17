import os
import json
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import re

def find_header_base_position(data):
    """
    動的にヘッダーの基準位置を検出
    1. 「項目」とある行を検索
    2. 「返礼品コード」がある列を検索
    """
    header_row_index = None
    header_col_index = None
    
    # 「項目」とある行を検索
    for i, row in enumerate(data):
        if len(row) >= 2 and row[1] == "項目":
            header_row_index = i
            break
    
    if header_row_index is None:
        return None, None
    
    # 「返礼品コード」がある列を検索（ヘッダー行内で）
    for j, cell in enumerate(data[header_row_index]):
        if cell and "返礼品コード" in str(cell):
            header_col_index = j
            break
    
    return header_row_index, header_col_index

def extract_hierarchical_headers(data, header_row_index, header_col_start):
    """
    階層的ヘッダー構造を解析
    メインヘッダーとサブヘッダーの関係を動的に抽出
    """
    if header_row_index is None or header_col_start is None:
        return []
    
    main_row = data[header_row_index]
    sub_row = data[header_row_index + 1] if header_row_index + 1 < len(data) else []
    
    hierarchical_headers = []
    
    # header_col_start から最後まで解析
    max_col = max(len(main_row), len(sub_row))
    
    for col_idx in range(header_col_start, max_col):
        main_header = main_row[col_idx] if col_idx < len(main_row) else None
        sub_header = sub_row[col_idx] if col_idx < len(sub_row) else None
        
        # ヘッダー名を生成
        header_info = create_unique_header_name(main_header, sub_header, col_idx, header_col_start)
        hierarchical_headers.append({
            'column_index': col_idx + 1,  # 1-based
            'main_header': main_header,
            'sub_header': sub_header,
            'unique_name': header_info['unique_name'],
            'standard_name': header_info['standard_name']
        })
    
    return hierarchical_headers

def create_unique_header_name(main_header, sub_header, col_idx, header_col_start):
    """
    ユニークなヘッダー名を生成
    空のヘッダーも「PAT-空N」形式で管理
    """
    # 基本的な正規化
    main_clean = normalize_text(main_header) if main_header else ""
    sub_clean = normalize_text(sub_header) if sub_header else ""
    
    # 標準名への変換ルール
    standard_mapping = {
        "返礼品コード": "返礼品コード",
        "ご記入日": "ご記入日",
        "事業者様名": "事業者様名",
        "事業者様TEL": "事業者様TEL",
        "商品名": "商品名",
        "産地": "産地",
        "内容量": "内容量",
        "発送温度帯": "発送温度帯",
        "保存方法": "保存方法",
        "リードタイム": "リードタイム",
    }
    
    # サブヘッダーの重要項目
    sub_mapping = {
        "発送元名称": "発送元名称",
        "住所": "発送元住所", 
        "TEL": "発送元TEL",
        "ご担当者様": "ご担当者様",
        "必須": "",  # 必須表示は無視
        "任意": "",  # 任意表示は無視
    }
    
    # 標準名の決定
    standard_name = None
    
    # 1. メインヘッダーでの完全一致
    if main_clean in standard_mapping:
        standard_name = standard_mapping[main_clean]
    
    # 2. サブヘッダーでの完全一致
    elif sub_clean in sub_mapping and sub_mapping[sub_clean]:
        standard_name = sub_mapping[sub_clean]
    
    # 3. メインヘッダーでの部分一致
    elif main_clean:
        for key, value in standard_mapping.items():
            if key in main_clean or main_clean in key:
                standard_name = value
                break
    
    # 4. サブヘッダーでの部分一致
    elif sub_clean:
        for key, value in sub_mapping.items():
            if key in sub_clean or sub_clean in key:
                if value:  # 空文字でない場合のみ
                    standard_name = value
                break
    
    # ユニーク名の生成
    if main_clean and sub_clean and sub_clean not in ["必須", "任意"]:
        unique_name = f"{main_clean}:{sub_clean}"
    elif main_clean:
        unique_name = main_clean
    elif sub_clean and sub_clean not in ["必須", "任意"]:
        unique_name = sub_clean
    else:
        # 空の場合は列位置で識別
        col_relative = col_idx - header_col_start + 1
        unique_name = f"空列{col_relative}"
    
    return {
        'unique_name': unique_name,
        'standard_name': standard_name or unique_name
    }

def normalize_text(text):
    """
    テキストの正規化
    """
    if not text:
        return ""
    
    text = str(text).strip()
    # 改行、連続空白を除去
    text = re.sub(r'\s+', '', text)
    # 特殊文字を除去
    text = re.sub(r'[^\w\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FAF]', '', text)
    
    return text

def analyze_file_structure(file_path):
    """
    ファイルの構造を動的に解析
    """
    file_name = os.path.basename(file_path)
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        max_row = ws.max_row
        max_col = ws.max_column

        # ワークシート全体の値を2次元リストにコピー
        data = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            data.append([cell.value for cell in row])

        # 動的にヘッダー基準位置を検出
        header_row_index, header_col_start = find_header_base_position(data)
        
        if header_row_index is None or header_col_start is None:
            print(f"  ⚠️ ヘッダー基準位置が見つかりません: {file_name}")
            return None
        
        print(f"  📍 {file_name}: ヘッダー基準位置 行{header_row_index + 1}, 列{header_col_start + 1}")
        
        # 階層的ヘッダーを抽出
        hierarchical_headers = extract_hierarchical_headers(data, header_row_index, header_col_start)
        
        # サンプルデータを取得
        sample_data = {}
        for header_info in hierarchical_headers:
            col_idx = header_info['column_index'] - 1  # 0-based
            samples = []
            for row_idx in range(header_row_index + 2, min(header_row_index + 6, len(data))):
                if row_idx < len(data) and col_idx < len(data[row_idx]):
                    val = data[row_idx][col_idx]
                    if val is not None and str(val).strip():
                        samples.append(str(val).strip())
            sample_data[header_info['unique_name']] = samples[:3]
        
        return {
            'file_name': file_name,
            'header_row_index': header_row_index,
            'header_col_start': header_col_start,
            'hierarchical_headers': hierarchical_headers,
            'sample_data': sample_data
        }
        
    except Exception as e:
        print(f"  ❌ 分析エラー: {file_name} → {e}")
        return None

def create_dynamic_mapping_config(file_structures):
    """
    動的解析結果から統合マッピング設定を作成
    """
    mapping_config = {}
    all_standard_names = set()
    
    # 全ファイルの標準名を収集
    for structure in file_structures:
        for header_info in structure['hierarchical_headers']:
            all_standard_names.add(header_info['standard_name'])
    
    all_standard_names = sorted(list(all_standard_names))
    
    # 各ファイルのマッピングを作成
    for structure in file_structures:
        file_mapping = {}
        
        for header_info in structure['hierarchical_headers']:
            standard_name = header_info['standard_name']
            column_index = header_info['column_index']
            
            # 同じ標準名が複数ある場合はユニーク名で区別
            if standard_name in file_mapping:
                # 既に存在する場合はユニーク名を使用
                unique_key = f"{standard_name}_{header_info['unique_name']}"
                file_mapping[unique_key] = column_index
            else:
                file_mapping[standard_name] = column_index
        
        mapping_config[structure['file_name']] = file_mapping
    
    return mapping_config, all_standard_names

def process_file_with_dynamic_mapping(file_path, file_mapping, all_standard_names, master_headers):
    """
    動的マッピングを使用してファイルを処理
    """
    file_name = os.path.basename(file_path)
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        max_row = ws.max_row
        max_col = ws.max_column

        # ワークシート全体の値を2次元リストにコピー
        data = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            data.append([cell.value for cell in row])

        # ヘッダー基準位置を再検出
        header_row_index, header_col_start = find_header_base_position(data)
        
        if header_row_index is None:
            print(f"  ⚠️ ヘッダー行が見つかりません: {file_name}")
            return [], master_headers

        # マスターヘッダーを更新
        for standard_name in all_standard_names:
            if standard_name not in master_headers:
                master_headers.append(standard_name)

        # データ行処理
        output_rows = []
        for row in data[header_row_index + 2:]:  # ヘッダー行+サブヘッダー行の次から
            new_row = []
            
            # A列: ファイル名
            existing_file_name = row[0] if len(row) > 0 and row[0] is not None else file_name
            new_row.append(existing_file_name)
            
            # B列: 項目値
            value_B = row[1] if len(row) > 1 else None
            new_row.append(value_B)
            
            # C列以降: 動的マッピングに基づいてデータを配置
            for standard_name in master_headers:
                value = None
                
                # このファイルでこの標準名に対応する列を探す
                if standard_name in file_mapping:
                    column_index = file_mapping[standard_name]
                    actual_index = column_index - 1  # 1-based から 0-based
                    if actual_index < len(row):
                        value = row[actual_index]
                
                new_row.append(value)
            
            output_rows.append(new_row)

        print(f"  ✅ 動的処理完了: {file_name} (行数: {len(output_rows)})")
        return output_rows, master_headers
        
    except Exception as e:
        print(f"  ❌ 処理エラー: {file_name} → {e}")
        return [], master_headers

def test_dynamic_header_mapping(municipality_name="不整合テスト自治体v3"):
    """
    動的ヘッダーマッピングのテスト
    """
    base_dir = os.path.join(
        r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase3\HARV',
        municipality_name
    )
    
    if not os.path.exists(base_dir):
        print(f"指定フォルダが存在しません: {base_dir}")
        return
    
    # 対象ファイルを取得（重複正規化ファイルは除外）
    files = [f for f in os.listdir(base_dir) 
             if f.startswith("PAT") and f.endswith("_normalized.xlsx") 
             and not f.endswith("_normalized_normalized.xlsx")]
    
    print(f"=== 動的ヘッダーマッピングテスト ===")
    print(f"対象ファイル: {len(files)}個")
    
    # STEP 1: 各ファイルの構造を動的解析
    print(f"\n=== STEP 1: ファイル構造解析 ===")
    file_structures = []
    
    for file in sorted(files):
        file_path = os.path.join(base_dir, file)
        print(f"\n--- {file} ---")
        
        structure = analyze_file_structure(file_path)
        if structure:
            file_structures.append(structure)
            
            # ヘッダー構造を表示
            print(f"  📋 検出ヘッダー数: {len(structure['hierarchical_headers'])}")
            for i, header_info in enumerate(structure['hierarchical_headers'][:10]):  # 最初の10個のみ表示
                print(f"    {header_info['column_index']:2d}列目: {header_info['unique_name']} → {header_info['standard_name']}")
            
            if len(structure['hierarchical_headers']) > 10:
                print(f"    ... 他 {len(structure['hierarchical_headers']) - 10} 個")
    
    # STEP 2: 動的マッピング設定を作成
    print(f"\n=== STEP 2: 動的マッピング設定作成 ===")
    mapping_config, all_standard_names = create_dynamic_mapping_config(file_structures)
    
    # マッピング設定をJSONで保存
    dynamic_mapping_path = os.path.join(base_dir, "dynamic_mapping_config.json")
    with open(dynamic_mapping_path, 'w', encoding='utf-8') as f:
        json.dump(mapping_config, f, ensure_ascii=False, indent=2)
    print(f"📁 動的マッピング設定保存: {dynamic_mapping_path}")
    
    print(f"📊 統合標準名数: {len(all_standard_names)}")
    
    # STEP 3: 動的マッピングで統合処理
    print(f"\n=== STEP 3: 動的マッピング統合処理 ===")
    
    master_headers = []
    master_data_rows = []
    
    for file in sorted(files):
        file_path = os.path.join(base_dir, file)
        file_mapping = mapping_config.get(file, {})
        
        print(f"\n--- {file} ---")
        rows, master_headers = process_file_with_dynamic_mapping(
            file_path, file_mapping, all_standard_names, master_headers
        )
        master_data_rows.extend(rows)
    
    # STEP 4: 統合結果を出力
    print(f"\n=== STEP 4: 結果出力 ===")
    
    dynamic_wb = Workbook()
    dynamic_ws = dynamic_wb.active
    dynamic_ws.title = "dynamic_mapping_integration"
    
    # ヘッダー設定
    dynamic_ws.cell(row=1, column=1, value="ファイル名")
    dynamic_ws.cell(row=1, column=2, value="項目")
    for i, header in enumerate(master_headers, start=3):
        dynamic_ws.cell(row=1, column=i, value=header)
    
    # データ設定
    current_row = 2
    for row in master_data_rows:
        for j, value in enumerate(row, start=1):
            dynamic_ws.cell(row=current_row, column=j, value=value)
        current_row += 1
    
    dynamic_output_path = os.path.join(base_dir, "dynamic_mapping_integration.xlsx")
    dynamic_wb.save(dynamic_output_path)
    
    print(f"📄 動的マッピング統合結果: {dynamic_output_path}")
    print(f"📊 統合ヘッダー数: {len(master_headers)}")
    print(f"📊 統合データ行数: {len(master_data_rows)}")
    
    # STEP 5: 詳細分析結果を出力
    print(f"\n=== STEP 5: 詳細分析 ===")
    
    analysis_path = os.path.join(base_dir, "header_analysis_report.json")
    analysis_report = {
        'file_structures': file_structures,
        'mapping_config': mapping_config,
        'all_standard_names': all_standard_names,
        'total_files': len(files),
        'total_headers': len(master_headers),
        'total_rows': len(master_data_rows)
    }
    
    with open(analysis_path, 'w', encoding='utf-8') as f:
        json.dump(analysis_report, f, ensure_ascii=False, indent=2)
    
    print(f"📁 詳細分析レポート: {analysis_path}")
    
    # 統合後のマスターヘッダー一覧
    print(f"\n=== 統合後のマスターヘッダー一覧 ===")
    for i, header in enumerate(master_headers, 1):
        print(f"  {i:2d}. {header}")
    
    print(f"\n✅ 動的ヘッダーマッピングテスト完了")
    
    return {
        'file_structures': file_structures,
        'mapping_config': mapping_config,
        'master_headers': master_headers,
        'total_rows': len(master_data_rows),
        'output_path': dynamic_output_path
    }

if __name__ == "__main__":
    result = test_dynamic_header_mapping()
