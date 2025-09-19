#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
カラムズレ問題調査スクリプト

問題：
- 商品名のところに担当者名が入る
- 会社名が内容量のところに入る
- 同じカラムの場合は一致させるルールで不整合が発生

調査対象：
- 2023-08-03_返礼品登録シート_株式会社ハマダ(赤身スライス900g定期便)（市確認）.xlsx
"""

import os
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import json
from datetime import datetime

# 調査対象ファイル
INVESTIGATION_TARGET = r"G:\共有ドライブ\k_40100_福岡県_北九州市_01\k_北九州市\99_共通\99_資料\02_返礼品関係\02_返礼品シート\企業名\h_株式会社ハマダ\2023-08-03_返礼品登録シート_株式会社ハマダ(赤身スライス900g定期便)（市確認）.xlsx"

# テスト用フォルダのファイル（比較用）
TEST_FOLDER = r"G:\共有ドライブ\★OD\99_商品管理\不整合データ\h_株式会社ハマダ"

def analyze_file_structure(file_path):
    """
    ファイルの構造を詳細に分析
    """
    print(f"\n=== ファイル構造分析: {os.path.basename(file_path)} ===")
    
    try:
        # data_only=Trueで値のみ取得
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"シート名: {ws.title}")
        print(f"最大行数: {ws.max_row}")
        print(f"最大列数: {ws.max_column}")
        
        # 結合セル情報
        merged_ranges = list(ws.merged_cells.ranges)
        print(f"結合セル数: {len(merged_ranges)}")
        if merged_ranges:
            print("結合セル一覧（最初の10個）:")
            for i, merged_range in enumerate(merged_ranges[:10]):
                print(f"  {i+1}: {merged_range}")
        
        # ヘッダー候補の検索
        header_candidates = find_header_candidates(ws)
        print(f"\nヘッダー候補:")
        for candidate in header_candidates:
            print(f"  行{candidate['row']}, 列{candidate['col']}: '{candidate['value']}'")
        
        # 特定のキーワードを含むセルを検索
        keywords = ["返礼品コード", "商品名", "担当者", "会社名", "内容量", "項目", "No."]
        keyword_locations = find_keyword_locations(ws, keywords)
        
        print(f"\nキーワード位置:")
        for keyword, locations in keyword_locations.items():
            if locations:
                print(f"  '{keyword}': {locations}")
        
        # データ範囲の特定
        data_range = identify_data_range(ws)
        print(f"\nデータ範囲: {data_range}")
        
        # 最初の10行×10列のデータサンプル
        print(f"\nデータサンプル（最初の10行×10列）:")
        sample_data = extract_sample_data(ws, 10, 10)
        for i, row in enumerate(sample_data, 1):
            print(f"  行{i}: {row}")
            
        return {
            'file_path': file_path,
            'sheet_name': ws.title,
            'max_row': ws.max_row,
            'max_col': ws.max_column,
            'merged_ranges': [str(r) for r in merged_ranges],
            'header_candidates': header_candidates,
            'keyword_locations': keyword_locations,
            'data_range': data_range,
            'sample_data': sample_data
        }
        
    except Exception as e:
        print(f"エラー: ファイル分析に失敗 - {e}")
        return None

def find_header_candidates(ws):
    """
    ヘッダー候補を検索
    """
    candidates = []
    
    # 最初の20行を検索
    for row in range(1, min(21, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 50)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                # ヘッダーらしいキーワード
                header_keywords = ["返礼品コード", "商品名", "項目", "No.", "必須", "任意"]
                if any(keyword in cell_value for keyword in header_keywords):
                    candidates.append({
                        'row': row,
                        'col': col,
                        'value': cell_value,
                        'coordinate': ws.cell(row=row, column=col).coordinate
                    })
    
    return candidates

def find_keyword_locations(ws, keywords):
    """
    特定のキーワードの位置を検索
    """
    locations = {keyword: [] for keyword in keywords}
    
    for row in range(1, min(ws.max_row + 1, 100)):  # 最初の100行を検索
        for col in range(1, min(ws.max_column + 1, 50)):  # 最初の50列を検索
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value)
                for keyword in keywords:
                    if keyword in cell_str:
                        locations[keyword].append({
                            'row': row,
                            'col': col,
                            'coordinate': ws.cell(row=row, column=col).coordinate,
                            'value': cell_str
                        })
    
    return locations

def identify_data_range(ws):
    """
    実際のデータ範囲を特定
    """
    # 実際にデータが存在する範囲を特定
    min_row, max_row = None, None
    min_col, max_col = None, None
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                if min_row is None:
                    min_row = row
                max_row = row
                if min_col is None or col < min_col:
                    min_col = col
                if max_col is None or col > max_col:
                    max_col = col
    
    return {
        'min_row': min_row,
        'max_row': max_row,
        'min_col': min_col,
        'max_col': max_col
    }

def extract_sample_data(ws, max_rows=10, max_cols=10):
    """
    サンプルデータを抽出
    """
    sample_data = []
    
    for row in range(1, min(max_rows + 1, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(max_cols + 1, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(cell_value)
        sample_data.append(row_data)
    
    return sample_data

def compare_phase_outputs():
    """
    Phase処理の各段階での出力を比較
    """
    print("\n=== Phase処理出力比較 ===")
    
    municipality = "2025-09-09_福岡県北九州市"
    
    # Phase1出力
    phase1_pattern_file = f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase1\\HARV\\{municipality}\\{municipality}_パターン一覧.xlsx"
    phase1_file_pattern = f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase1\\HARV\\{municipality}\\{municipality}_ファイル別パターン.xlsx"
    
    # Phase2出力
    phase2_file = f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase2\\HARV\\{municipality}\\{municipality}_パターン一覧_Phase2.xlsx"
    
    # Phase3出力
    phase3_dir = f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase3\\HARV\\{municipality}"
    
    files_to_check = [
        ("Phase1パターン一覧", phase1_pattern_file),
        ("Phase1ファイル別パターン", phase1_file_pattern),
        ("Phase2パターン一覧", phase2_file),
    ]
    
    for name, file_path in files_to_check:
        if os.path.exists(file_path):
            print(f"\n--- {name} ---")
            try:
                df = pd.read_excel(file_path)
                print(f"形状: {df.shape}")
                print(f"列名: {list(df.columns)}")
                print("最初の3行:")
                print(df.head(3).to_string())
            except Exception as e:
                print(f"読み込みエラー: {e}")
        else:
            print(f"\n--- {name} ---")
            print("ファイルが存在しません")
    
    # Phase3のPATファイルをチェック
    if os.path.exists(phase3_dir):
        print(f"\n--- Phase3 PATファイル ---")
        pat_files = [f for f in os.listdir(phase3_dir) if f.startswith("PAT") and f.endswith(".xlsx")]
        print(f"PATファイル数: {len(pat_files)}")
        for pat_file in pat_files[:5]:  # 最初の5個
            print(f"  {pat_file}")

def trace_hamada_file_processing():
    """
    ハマダファイルの処理過程を追跡
    """
    print("\n=== ハマダファイル処理追跡 ===")
    
    target_file = "2023-08-03_返礼品登録シート_株式会社ハマダ(赤身スライス900g定期便)（市確認）.xlsx"
    
    # Phase1でのパターン検出結果を確認
    municipality = "2025-09-09_福岡県北九州市"
    phase1_file_pattern = f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase1\\HARV\\{municipality}\\{municipality}_ファイル別パターン.xlsx"
    
    if os.path.exists(phase1_file_pattern):
        try:
            df = pd.read_excel(phase1_file_pattern)
            hamada_records = df[df['ファイル名'].str.contains('ハマダ', na=False)]
            print("ハマダ関連ファイルのパターン:")
            for _, row in hamada_records.iterrows():
                print(f"  {row['ファイル名']} -> {row['パターン名']}")
            
            # 問題のファイルを特定
            target_records = df[df['ファイル名'] == target_file]
            if not target_records.empty:
                target_pattern = target_records.iloc[0]['パターン名']
                print(f"\n問題のファイルのパターン: {target_pattern}")
                
                # そのパターンの定義を確認
                phase2_file = f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase2\\HARV\\{municipality}\\{municipality}_パターン一覧_Phase2.xlsx"
                if os.path.exists(phase2_file):
                    df2 = pd.read_excel(phase2_file, header=None)
                    pattern_rows = df2[df2[0] == target_pattern]
                    if not pattern_rows.empty:
                        print(f"パターン定義:")
                        for _, row in pattern_rows.iterrows():
                            print(f"  {list(row)}")
                
        except Exception as e:
            print(f"Phase1ファイル読み込みエラー: {e}")

def analyze_column_mapping_logic():
    """
    カラムマッピングロジックの分析
    """
    print("\n=== カラムマッピングロジック分析 ===")
    
    # merge.pyの関連部分を分析
    print("merge.pyのPhase5処理における問題点:")
    print("1. find_header_base_position関数:")
    print("   - '項目'とある行を検索")
    print("   - '返礼品コード'がある列を検索")
    print("   - この検索が曖昧な場合、間違った基準位置を設定する可能性")
    
    print("\n2. extract_dynamic_headers関数:")
    print("   - main_row（メインヘッダー行）とsub_row（サブヘッダー行）を取得")
    print("   - create_hierarchical_header関数で階層的ヘッダー名を生成")
    print("   - ヘッダー名の正規化処理で情報が失われる可能性")
    
    print("\n3. process_file関数:")
    print("   - header_to_column マッピングでデータを配置")
    print("   - 同一ヘッダー名で異なる意味のデータが統合される可能性")
    
    print("\n4. 特定の修正コード（734-737行）:")
    print("   - PAT0001_normalized.xlsx の商品名列から '河瀨' を除去")
    print("   - この種の個別対応が他のファイルに影響を与える可能性")

def investigate_specific_misalignment():
    """
    具体的な不整合パターンを調査
    """
    print("\n=== 具体的な不整合パターン調査 ===")
    
    # 仮説1: ヘッダー検出の失敗
    print("仮説1: ヘッダー検出の失敗")
    print("- '項目'や'返礼品コード'が複数箇所にある場合の誤検出")
    print("- 結合セルによる位置のずれ")
    
    # 仮説2: 階層ヘッダーの処理問題
    print("\n仮説2: 階層ヘッダーの処理問題")
    print("- メインヘッダーとサブヘッダーの組み合わせが不正確")
    print("- 正規化処理での情報の欠落")
    
    # 仮説3: パターンマッピングの問題
    print("\n仮説3: パターンマッピングの問題")
    print("- 同一パターン名で異なる構造のファイルが存在")
    print("- マスターヘッダーの統合時の順序問題")
    
    # 仮説4: データ統合時の問題
    print("\n仮説4: データ統合時の問題")
    print("- header_to_column マッピングの不整合")
    print("- 列インデックスの計算ミス")

def main():
    """
    メイン調査処理
    """
    print("=== カラムズレ問題調査開始 ===")
    print(f"調査開始時刻: {datetime.now()}")
    
    # 1. 問題のファイルの構造分析
    if os.path.exists(INVESTIGATION_TARGET):
        target_analysis = analyze_file_structure(INVESTIGATION_TARGET)
    else:
        print(f"調査対象ファイルが見つかりません: {INVESTIGATION_TARGET}")
        target_analysis = None
    
    # 2. テスト用フォルダの同じファイルと比較
    test_file = os.path.join(TEST_FOLDER, "2023-08-03_返礼品登録シート_株式会社ハマダ(赤身スライス900g定期便)（市確認）.xlsx")
    if os.path.exists(test_file):
        test_analysis = analyze_file_structure(test_file)
    else:
        print(f"テスト用ファイルが見つかりません: {test_file}")
        test_analysis = None
    
    # 3. Phase処理の出力比較
    compare_phase_outputs()
    
    # 4. ハマダファイルの処理追跡
    trace_hamada_file_processing()
    
    # 5. カラムマッピングロジック分析
    analyze_column_mapping_logic()
    
    # 6. 具体的な不整合パターン調査
    investigate_specific_misalignment()
    
    # 7. 結果の保存
    results = {
        'investigation_time': str(datetime.now()),
        'target_file_analysis': target_analysis,
        'test_file_analysis': test_analysis
    }
    
    output_file = "debug/column_misalignment_investigation_results.json"
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2, default=str)
        print(f"\n調査結果を保存しました: {output_file}")
    except Exception as e:
        print(f"結果保存エラー: {e}")
    
    print("\n=== 調査完了 ===")

if __name__ == "__main__":
    main()
