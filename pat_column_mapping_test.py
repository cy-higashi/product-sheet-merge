import os
import json
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import re

def create_column_mapping_config():
    """
    PAT別の列マッピング設定を作成
    検出された問題を基に、各PATでの実際のヘッダー位置をマッピング
    """
    mapping_config = {
        "PAT0001_normalized.xlsx": {
            "返礼品コード": 3,
            "ご記入日": 4,  # 空の場合あり
            "事業者様名": 5,  # 実際は「ご記入日」だが事業者様名として扱う
            "事業者様TEL": 6,  # 実際は「事業者様名」
            "発送元名称": 7,  # 実際は「事業者様TEL」
            "発送元住所": 8,  # 「返礼品発送元情報」
            "発送元TEL": 9,  # 「返礼品発送元情報」
            "ご担当者様": 10, # 「返礼品発送元情報」
            "商品名": 11,     # 実際は「ご担当者様」だが商品名として扱う
            "産地": 12,       # 実際の「産地」
            "内容量": 14,     # 実際は「生産者・製造者・加工元」だが内容量として扱う
            "発送温度帯": 20,
            "保存方法": 21,   # 実際は「発送温度帯」だが保存方法として扱う
            "受付期間": 22,   # 実際は「保存方法」だが受付期間として扱う
            "発送期間": 24,   # 実際は「受付終了」だが発送期間として扱う
            "リードタイム": 26, # 実際は「発送終了」だがリードタイムとして扱う
        },
        "PAT0002_normalized.xlsx": {
            "返礼品コード": 3,
            "ご記入日": 4,
            "事業者様名": 5,
            "事業者様TEL": 6,
            "発送元名称": 7,
            "発送元住所": 8,
            "発送元TEL": 9,
            "ご担当者様": 10,
            "商品名": 11,
            "商品名(伝票記載用)": 12,
            "内容量": 24,
            "発送温度帯": 31,
            "保存方法": 32,
            "受付期間": 33,
            "発送期間": 35,
            "リードタイム": 37,
        },
        "PAT0003_normalized.xlsx": {
            "返礼品コード": 3,
            "返礼品コード_2": 4,  # 重複
            "ご記入日": 5,
            "事業者様名": 6,
            "事業者様TEL": 7,
            "発送元名称": 8,
            "発送元住所": 9,
            "発送元TEL": 10,
            "ご担当者様": 11,
            "商品名": 12,
            "内容量": 15,
            "発送温度帯": 21,
            "保存方法": 22,
            "受付期間": 23,
            "発送期間": 25,
            "リードタイム": 27,
        },
        "PAT0004_normalized.xlsx": {
            "返礼品コード": 3,
            "ご記入日": 4,
            "事業者様名": 5,
            "事業者様TEL": 6,
            "発送元名称": 7,
            "発送元住所": 8,
            "発送元TEL": 9,
            "ご担当者様": 10,
            "商品名": 11,
            "商品名(伝票記載用)": 12,
            "内容量": 24,
            "発送温度帯": 31,
            "保存方法": 32,
            "受付期間": 33,
            "発送期間": 35,
            "リードタイム": 37,
        },
        "PAT0005_normalized.xlsx": {
            "返礼品コード": 3,
            "ご記入日": 4,
            "事業者様名": 5,
            "事業者様TEL": 6,
            "発送元名称": 7,
            "発送元住所": 8,
            "発送元TEL": 9,
            "ご担当者様": 10,
            "商品名": 11,
            "産地": 12,
            "内容量": 14,
            "発送温度帯": 20,
            "保存方法": 21,
            "受付期間": 22,
            "発送期間": 24,
            "リードタイム": 26,
        },
        "PAT0006_normalized.xlsx": {
            "返礼品コード": 3,
            "ご記入日": 4,
            "事業者様名": 5,
            "事業者様TEL": 6,
            "発送元名称": 7,
            "発送元住所": 8,
            "発送元TEL": 9,
            "ご担当者様": 10,
            "商品名": 11,
            "商品名(伝票記載用)": 12,
            "内容量": 24,
            "発送温度帯": 31,
            "保存方法": 32,
            "受付期間": 33,
            "発送期間": 35,
            "リードタイム": 37,
        }
    }
    return mapping_config

def normalize_header_name(header):
    """
    ヘッダー名を完全一意化する
    """
    if not header:
        return ""
    
    # 基本的な正規化
    normalized = str(header).strip()
    
    # 改行、連続空白を除去
    normalized = re.sub(r'\s+', ' ', normalized)
    
    # 一意化のための変換ルール
    mapping_rules = {
        # 事業者関連
        "事業者様名": "事業者様名",
        "事業者名": "事業者様名",
        "事業者様TEL": "事業者様TEL",
        "事業者TEL": "事業者様TEL",
        
        # 発送元関連
        "発送元名称": "発送元名称",
        "発送元": "発送元名称",
        "発送元住所": "発送元住所",
        "住所": "発送元住所",
        "発送元TEL": "発送元TEL",
        "TEL": "発送元TEL",
        "ご担当者様": "ご担当者様",
        "担当者": "ご担当者様",
        
        # 商品関連
        "商品名": "商品名",
        "商品名(伝票記載用)": "商品名(伝票記載用)",
        "商品名（伝票記載用）": "商品名(伝票記載用)",
        "産地": "産地",
        "内容量": "内容量",
        
        # 日付・期間関連
        "ご記入日": "ご記入日",
        "記入日": "ご記入日",
        "受付期間": "受付期間",
        "受付開始": "受付期間",
        "発送期間": "発送期間",
        "発送開始": "発送期間",
        "リードタイム": "リードタイム",
        
        # 保存・配送関連
        "発送温度帯": "発送温度帯",
        "温度帯": "発送温度帯",
        "保存方法": "保存方法",
        
        # 返礼品コード
        "返礼品コード": "返礼品コード",
    }
    
    # 完全一致での変換
    if normalized in mapping_rules:
        return mapping_rules[normalized]
    
    # 部分一致での変換
    for pattern, standard in mapping_rules.items():
        if pattern in normalized:
            return standard
    
    # 変換されなかった場合はそのまま返す
    return normalized

def process_file_with_mapping(file_path, mapping_config, master_headers):
    """
    列マッピング設定を使用してファイルを処理
    """
    file_name = os.path.basename(file_path)
    
    if file_name not in mapping_config:
        print(f"  ⚠️ マッピング設定なし: {file_name}")
        return [], master_headers
    
    column_mapping = mapping_config[file_name]
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        max_row = ws.max_row
        max_col = ws.max_column

        # ワークシート全体の値を2次元リストにコピー
        data = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            data.append([cell.value for cell in row])

        # ヘッダー行（B列が「項目」）を検索
        header_row_index = None
        for i, row in enumerate(data):
            if len(row) >= 2 and row[1] == "項目":
                header_row_index = i
                break

        if header_row_index is None:
            print(f"  ⚠️ 「項目」行が見つかりません: {file_name}")
            return [], master_headers

        # マッピング設定に基づいて標準化されたヘッダーを生成
        standardized_headers = []
        for standard_name, column_index in column_mapping.items():
            normalized_name = normalize_header_name(standard_name)
            if normalized_name not in master_headers:
                master_headers.append(normalized_name)
            standardized_headers.append(normalized_name)

        # データ行処理
        output_rows = []
        for row in data[header_row_index+1:]:
            new_row = []
            
            # A列: ファイル名
            existing_file_name = row[0] if len(row) > 0 and row[0] is not None else file_name
            new_row.append(existing_file_name)
            
            # B列: 項目値
            value_B = row[1] if len(row) > 1 else None
            new_row.append(value_B)
            
            # C列以降: マッピング設定に基づいてデータを配置
            for header in master_headers:
                value = None
                # このファイルでこのヘッダーに対応するマッピングを探す
                for standard_name, column_index in column_mapping.items():
                    if normalize_header_name(standard_name) == header:
                        # 1-based から 0-based に変換
                        actual_index = column_index - 1
                        if actual_index < len(row):
                            value = row[actual_index]
                        break
                new_row.append(value)
            
            output_rows.append(new_row)

        print(f"  ✅ マッピング処理完了: {file_name} (行数: {len(output_rows)}, 標準ヘッダー: {len(standardized_headers)})")
        return output_rows, master_headers
        
    except Exception as e:
        print(f"  ❌ エラー: {file_name} → {e}")
        return [], master_headers

def validate_mapping_accuracy(base_dir, mapping_config):
    """
    マッピング設定の精度を検証
    """
    print("\n=== マッピング精度検証 ===")
    
    validation_results = {}
    
    for file_name, column_mapping in mapping_config.items():
        file_path = os.path.join(base_dir, file_name)
        if not os.path.exists(file_path):
            continue
            
        print(f"\n--- {file_name} の検証 ---")
        
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # データを取得
            data = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                data.append([cell.value for cell in row])

            # ヘッダー行を検索
            header_row_index = None
            for i, row in enumerate(data):
                if len(row) >= 2 and row[1] == "項目":
                    header_row_index = i
                    break

            if header_row_index is None:
                continue

            # マッピングの精度をチェック
            accurate_mappings = 0
            total_mappings = 0
            
            for standard_name, column_index in column_mapping.items():
                total_mappings += 1
                actual_index = column_index - 1  # 1-based から 0-based
                
                if actual_index < len(data[header_row_index]):
                    actual_header = data[header_row_index][actual_index]
                    
                    # サンプルデータを取得
                    sample_data = []
                    for row_idx in range(header_row_index + 1, min(header_row_index + 4, len(data))):
                        if row_idx < len(data) and actual_index < len(data[row_idx]):
                            val = data[row_idx][actual_index]
                            if val is not None and str(val).strip():
                                sample_data.append(str(val).strip())
                    
                    print(f"  {standard_name} → {column_index}列目")
                    print(f"    実際のヘッダー: {actual_header}")
                    print(f"    サンプルデータ: {sample_data[:2]}")
                    
                    # 精度判定（簡易版）
                    if is_mapping_accurate(standard_name, actual_header, sample_data):
                        accurate_mappings += 1
                        print(f"    ✅ 適切")
                    else:
                        print(f"    ⚠️ 要確認")
            
            accuracy = (accurate_mappings / total_mappings) * 100 if total_mappings > 0 else 0
            validation_results[file_name] = {
                'accuracy': accuracy,
                'accurate': accurate_mappings,
                'total': total_mappings
            }
            print(f"  📊 精度: {accuracy:.1f}% ({accurate_mappings}/{total_mappings})")
            
        except Exception as e:
            print(f"  ❌ 検証エラー: {e}")
    
    return validation_results

def is_mapping_accurate(standard_name, actual_header, sample_data):
    """
    マッピングが適切かどうかを判定
    """
    # 完全一致チェック
    if actual_header == standard_name:
        return True
    
    # 正規化後の一致チェック
    if normalize_header_name(actual_header) == normalize_header_name(standard_name):
        return True
    
    # サンプルデータによる内容一致チェック
    if standard_name in ["ご記入日"] and sample_data:
        # 日付データの形式チェック
        for sample in sample_data:
            if re.match(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}|^\d{5}$', sample):  # 日付またはExcel日付シリアル
                return True
    
    if standard_name in ["事業者様TEL", "発送元TEL"] and sample_data:
        # 電話番号の形式チェック
        for sample in sample_data:
            if re.match(r'[\d\-\(\)]+', sample):
                return True
    
    # その他の判定ロジックを追加可能
    return False

def test_column_mapping_integration(municipality_name="不整合テスト自治体v3"):
    """
    列マッピング設定による統合テスト
    """
    base_dir = os.path.join(
        r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase3\HARV',
        municipality_name
    )
    
    if not os.path.exists(base_dir):
        print(f"指定フォルダが存在しません: {base_dir}")
        return
    
    # 対象ファイルを取得（重複正規化ファイルは除外）
    files = [f for f in os.listdir(base_dir) 
             if f.startswith("PAT") and f.endswith("_normalized.xlsx") 
             and not f.endswith("_normalized_normalized.xlsx")]
    
    print(f"=== 列マッピング統合テスト ===")
    print(f"対象ファイル: {len(files)}個")
    
    # マッピング設定を作成
    mapping_config = create_column_mapping_config()
    
    # マッピング設定をJSONで保存
    mapping_path = os.path.join(base_dir, "column_mapping_config.json")
    with open(mapping_path, 'w', encoding='utf-8') as f:
        json.dump(mapping_config, f, ensure_ascii=False, indent=2)
    print(f"📁 マッピング設定保存: {mapping_path}")
    
    # マッピング精度を検証
    validation_results = validate_mapping_accuracy(base_dir, mapping_config)
    
    print(f"\n=== マッピング統合処理 ===")
    
    # マッピングベース統合処理
    master_headers = []
    master_data_rows = []
    
    for file in sorted(files):
        file_path = os.path.join(base_dir, file)
        print(f"\n--- {file} ---")
        
        rows, master_headers = process_file_with_mapping(file_path, mapping_config, master_headers)
        master_data_rows.extend(rows)
    
    # 統合結果を出力
    mapping_wb = Workbook()
    mapping_ws = mapping_wb.active
    mapping_ws.title = "mapping_based_integration"
    
    # ヘッダー設定
    mapping_ws.cell(row=1, column=1, value="ファイル名")
    mapping_ws.cell(row=1, column=2, value="項目")
    for i, header in enumerate(master_headers, start=3):
        mapping_ws.cell(row=1, column=i, value=header)
    
    # データ設定
    current_row = 2
    for row in master_data_rows:
        for j, value in enumerate(row, start=1):
            mapping_ws.cell(row=current_row, column=j, value=value)
        current_row += 1
    
    mapping_output_path = os.path.join(base_dir, "mapping_based_integration.xlsx")
    mapping_wb.save(mapping_output_path)
    
    print(f"\n=== 統合結果 ===")
    print(f"📄 マッピングベース統合結果: {mapping_output_path}")
    print(f"📊 統合ヘッダー数: {len(master_headers)}")
    print(f"📊 統合データ行数: {len(master_data_rows)}")
    
    # 検証サマリー
    print(f"\n=== マッピング精度サマリー ===")
    for file_name, result in validation_results.items():
        print(f"  {file_name}: {result['accuracy']:.1f}% ({result['accurate']}/{result['total']})")
    
    avg_accuracy = sum(r['accuracy'] for r in validation_results.values()) / len(validation_results) if validation_results else 0
    print(f"📈 平均精度: {avg_accuracy:.1f}%")
    
    # ヘッダー一覧を表示
    print(f"\n=== 統合後の標準ヘッダー一覧 ===")
    for i, header in enumerate(master_headers, 1):
        print(f"  {i:2d}. {header}")
    
    print(f"\n✅ 列マッピング統合テスト完了")
    
    return {
        'mapping_config': mapping_config,
        'validation_results': validation_results,
        'master_headers': master_headers,
        'total_rows': len(master_data_rows),
        'output_path': mapping_output_path
    }

if __name__ == "__main__":
    result = test_column_mapping_integration()
