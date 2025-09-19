import os
import re
from pathlib import Path
import time
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from copy import copy

# ===== 定数設定 =====
# 環境変数がなければデフォルト値を利用
MUNICIPALITY_NAME = os.getenv('MUNICIPALITY_NAME', '熊本市')
TARGET_PATH = os.getenv('TARGET_PATH', r'G:\共有ドライブ\★OD_管理者\データマネジメント部\DataOps\オペレーション\商品管理\test_data')

# 出力先ディレクトリ
PHASE1_OUTPUT_DIR = os.path.join(
    r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase1\HARV',
    MUNICIPALITY_NAME
)
PHASE2_OUTPUT_DIR = os.path.join(
    r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase2\HARV',
    MUNICIPALITY_NAME
)
PHASE3_OUTPUT_DIR = os.path.join(
    r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase3\HARV',
    MUNICIPALITY_NAME
)
LOG_FILE_PATH = os.path.join(PHASE1_OUTPUT_DIR, "execution_log.txt")

# 出力先ディレクトリが存在しなければ作成
for d in [PHASE1_OUTPUT_DIR, PHASE2_OUTPUT_DIR, PHASE3_OUTPUT_DIR]:
    if not os.path.exists(d):
        os.makedirs(d, exist_ok=True)

# ===== Phase1: パターン一覧とファイル別パターン作成 =====
def process_phase1(target_path, municipality_name, phase1_output_dir, log_file_path):
    with open(log_file_path, 'a', encoding='utf-8') as log_file:
        log_file.write(f"\n\n==== Phase1 実行開始: {datetime.now()} ====\nターゲットパス: {target_path}\n")
    
    target_path = Path(target_path)
    xlsx_files = []
    # サブフォルダ内の.xlsxファイルを収集
    for folder in target_path.iterdir():
        if folder.is_dir() and re.match(r'^[a-zA-Z0-9]', folder.name):
            for xlsx_file in folder.glob('*.xlsx'):
                xlsx_files.append((folder.name, xlsx_file))
    
    with open(log_file_path, 'a', encoding='utf-8') as log_file:
        log_file.write(f"Found {len(xlsx_files)} xlsx files in target path\n")
    
    output_data = []       # パターン定義用データ
    file_pattern_data = [] # ファイル別パターン情報
    pattern_counter = 0
    existing_patterns = {}

    def cache_merged_cells(sheet):
        merged_cells_cache = {}
        for merged_range in sheet.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cells_cache[(row, col)] = (merged_range.min_row, merged_range.min_col)
        return merged_cells_cache

    def get_merged_cell_value(sheet, cell, merged_cells_cache):
        if (cell.row, cell.column) in merged_cells_cache:
            min_row, min_col = merged_cells_cache[(cell.row, cell.column)]
            return sheet.cell(min_row, min_col).value
        return cell.value

    def get_right_column_value(sheet, row, column, merged_cells_cache):
        if column + 1 <= sheet.max_column:
            right_col_cell = sheet.cell(row, column + 1)
            right_value = get_merged_cell_value(sheet, right_col_cell, merged_cells_cache)
            if right_value:
                return f"+++{right_value}"
        return ""

    def get_values_until_last_data(sheet, start_cell, merged_cells_cache):
        values = []
        empty_count = 0
        max_empty_cells = 10
        for r in range(start_cell.row + 1, sheet.max_row + 1):
            cell = sheet.cell(r, start_cell.column)
            value = get_merged_cell_value(sheet, cell, merged_cells_cache)
            if value is None or value == "":
                empty_count += 1
            else:
                empty_count = 0
                if r > start_cell.row + 1 and value == get_merged_cell_value(sheet, sheet.cell(r - 1, start_cell.column), merged_cells_cache):
                    right_value = get_right_column_value(sheet, r, start_cell.column, merged_cells_cache)
                    value = f"{value}{right_value}"
            values.append(value)
            if empty_count >= max_empty_cells:
                break
        return [v for v in values if v is not None and v != ""]

    def find_existing_pattern(pattern_data):
        for pattern_name, existing_pattern in existing_patterns.items():
            if pattern_data == existing_pattern:
                return pattern_name
        return None

    for i, (folder_name, xlsx_file) in enumerate(xlsx_files):
        file_path = str(xlsx_file)
        file_name = xlsx_file.name
        try:
            print(f"Processing file: {file_name}")
            with open(log_file_path, 'a', encoding='utf-8') as log_file:
                log_file.write(f"\nProcessing file: {file_name}\n")
            try:
                workbook = load_workbook(file_path, data_only=True)
            except Exception as e:
                with open(log_file_path, 'a', encoding='utf-8') as log_file:
                    log_file.write(f"Failed to load workbook {file_name}: {e}\n")
                continue
            sheet = workbook.active
            merged_cells_cache = cache_merged_cells(sheet)
            pattern_found = False
            file_id = None  # ローカルファイルではIDは不要
            for row in sheet.iter_rows():
                for cell in row:
                    cell_value = get_merged_cell_value(sheet, cell, merged_cells_cache)
                    if cell_value == '返礼品コード':
                        all_values = get_values_until_last_data(sheet, cell, merged_cells_cache)
                        existing_pattern_name = find_existing_pattern(all_values)
                        if existing_pattern_name:
                            file_pattern_data.append([municipality_name, folder_name, file_name, existing_pattern_name, file_id])
                        else:
                            pattern_counter += 1
                            pattern_name = f"PAT{str(pattern_counter).zfill(4)}"
                            output_data.append([pattern_name, f"{cell.column_letter}{cell.row}"] + all_values)
                            file_pattern_data.append([municipality_name, folder_name, file_name, pattern_name, file_id])
                            existing_patterns[pattern_name] = all_values
                        pattern_found = True
                        break
                if pattern_found:
                    break
            if not pattern_found:
                file_pattern_data.append([municipality_name, folder_name, file_name, 'なし', file_id])
            with open(log_file_path, 'a', encoding='utf-8') as log_file:
                log_file.write(f"File {i+1}/{len(xlsx_files)} processed: {file_name}\n")
        except Exception as e:
            with open(log_file_path, 'a', encoding='utf-8') as log_file:
                log_file.write(f"Error processing {file_name}: {e}\n")
    
    if output_data:
        max_columns = max([len(row) for row in output_data])
        column_names = ['パターン名', 'A1形式'] + [f'列の値_{i}' for i in range(1, max_columns - 1)]
        output_df = pd.DataFrame(output_data, columns=column_names)
        file_pattern_df = pd.DataFrame(file_pattern_data, columns=['自治体', 'フォルダ名', 'ファイル名', 'パターン名', 'ファイルID'])
        output_path = os.path.join(phase1_output_dir, f"{municipality_name}_パターン一覧.xlsx")
        try:
            output_df.to_excel(output_path, index=False)
        except Exception as e:
            with open(log_file_path, 'a', encoding='utf-8') as log_file:
                log_file.write(f"Failed to save output Excel file: {e}\n")
        file_pattern_output_path = os.path.join(phase1_output_dir, f"{municipality_name}_ファイル別パターン.xlsx")
        try:
            file_pattern_df.to_excel(file_pattern_output_path, index=False)
        except Exception as e:
            with open(log_file_path, 'a', encoding='utf-8') as log_file:
                log_file.write(f"Failed to save file pattern Excel file: {e}\n")
        with open(log_file_path, 'a', encoding='utf-8') as log_file:
            log_file.write(f"\nResults saved to: {output_path}\nFile patterns saved to: {file_pattern_output_path}\n")
            log_file.write(f"==== Phase1 実行終了: {datetime.now()} ====\n\n")

# ===== Phase2: パターン一覧_Phase2.xlsx 作成 =====
def process_phase2(municipality_name, phase1_output_dir, phase2_output_dir):
    input_file = os.path.join(phase1_output_dir, f"{municipality_name}_パターン一覧.xlsx")
    output_file = os.path.join(phase2_output_dir, f"{municipality_name}_パターン一覧_Phase2.xlsx")
    df = pd.read_excel(input_file, sheet_name="Sheet1", header=None)
    df = df.fillna("")
    data = df.values.tolist()
    data = [row for row in data if str(row[0]).strip() != ""]
    last_row = len(data)
    if last_row == 0:
        raise ValueError("入力ファイルにデータがありません。")
    total_columns = len(data[0])
    unique_headers = {}
    header_order = []
    for col in range(total_columns - 1, 1, -1):
        for r in range(1, last_row):
            cell_value = str(data[r][col]).strip()
            if cell_value != "":
                if cell_value not in unique_headers:
                    unique_headers[cell_value] = True
                    header_order.append(cell_value)
    headers = header_order[::-1]
    new_data = []
    new_header_row = [data[0][0], data[0][1]] + headers
    new_data.append(new_header_row)
    for r in range(1, last_row):
        new_row = [data[r][0], data[r][1]]
        row_data = {}
        for cell in data[r]:
            cell_str = str(cell).strip()
            if cell_str != "":
                row_data[cell_str] = cell_str
        for h in headers:
            new_row.append(row_data.get(h, ""))
        new_data.append(new_row)
    for i, row in enumerate(new_data):
        if i == 0:
            row.insert(2, "返礼品コード")
        else:
            if isinstance(row[0], str) and row[0].startswith("PAT"):
                row.insert(2, "返礼品コード")
            else:
                row.insert(2, "")
    num_rows = len(new_data)
    num_cols = len(new_data[0])
    col_counts = []
    for col in range(num_cols):
        count = sum(1 for row in new_data[1:] if str(row[col]).strip() != "")
        col_counts.append((col, count))
    col_counts_sorted = sorted(col_counts, key=lambda x: x[1], reverse=True)
    reordered_data = []
    for row in new_data:
        new_row = [row[col_index] for col_index, _ in col_counts_sorted]
        reordered_data.append(new_row)
    output_df = pd.DataFrame(reordered_data)
    output_df.to_excel(output_file, index=False, header=False)
    print(f"Phase2 の処理が完了しました。\n出力ファイル: {output_file}")

# ===== Phase3: 各ファイルの転置処理 =====
def load_pattern_map(def_file, sheet_name="Sheet1"):
    try:
        df = pd.read_excel(def_file, sheet_name=sheet_name)
    except Exception as e:
        print(f"パターン定義シートの読み込みに失敗: {e}")
        return {}
    pattern_map = {}
    if df.empty:
        return pattern_map
    pattern_col = df.columns[0]
    output_cols = list(df.columns[1:])
    for idx, row in df.iterrows():
        pattern_name = str(row[pattern_col]).strip() if pd.notna(row[pattern_col]) else ""
        if not pattern_name:
            continue
        mapping = {}
        for col in output_cols:
            cell_val = row[col]
            if pd.notna(cell_val):
                keyword = str(cell_val).strip()
                if keyword.startswith("キーワード_"):
                    keyword = keyword.replace("キーワード_", "")
            else:
                keyword = ""
            mapping[col] = keyword
        pattern_map[pattern_name] = mapping
    return pattern_map

def load_file_list(list_file, sheet_name="Sheet1"):
    try:
        df = pd.read_excel(list_file, sheet_name=sheet_name)
    except Exception as e:
        print(f"ファイル一覧シートの読み込みに失敗: {e}")
        return None
    required_columns = ["自治体", "フォルダ名", "ファイル名", "パターン名"]
    for col in required_columns:
        if col not in df.columns:
            print(f"必要な列 '{col}' がファイル一覧にありません。")
            return None
    return df

def process_file_phase3(file_path, combined_ws, start_output_row):
    wb = load_workbook(file_path)
    ws = wb.active
    start_cell = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value == "No.":
                start_cell = cell
                break
        if start_cell:
            break
    if not start_cell:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value == "項目":
                    start_cell = cell
                    break
            if start_cell:
                break
    if not start_cell:
        raise ValueError(f'"No." または "項目" が {file_path} 内に見つかりませんでした。')
    start_row, start_col = start_cell.row, start_cell.column
    end_col_candidate = None
    if start_cell.value == "項目":
        for merged_range in ws.merged_cells.ranges:
            if start_cell.coordinate in merged_range:
                start_col = merged_range.min_col
                end_col_candidate = merged_range.max_col
                break
    end_row = start_row
    for r in range(start_row, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(start_col, ws.max_column + 1)):
            end_row = r
    end_col = start_col
    for c in range(start_col, ws.max_column + 1):
        if any(ws.cell(row=r, column=c).value is not None for r in range(start_row, end_row + 1)):
            end_col = c
    if end_col_candidate is not None and end_col_candidate > end_col:
        end_col = end_col_candidate
    orig_rows = end_row - start_row + 1
    orig_cols = end_col - start_col + 1
    for r_idx, r in enumerate(range(start_row, end_row + 1), start=1):
        for c_idx, c in enumerate(range(start_col, end_col + 1), start=1):
            orig_cell = ws.cell(row=r, column=c)
            final_row = start_output_row + (c_idx - 1)
            final_col = r_idx + 1
            new_cell = combined_ws.cell(row=final_row, column=final_col, value=orig_cell.value)
            if orig_cell.has_style:
                new_cell.font = copy(orig_cell.font)
                new_cell.border = copy(orig_cell.border)
                new_cell.fill = copy(orig_cell.fill)
                new_cell.number_format = copy(orig_cell.number_format)
                new_cell.protection = copy(orig_cell.protection)
                new_cell.alignment = copy(orig_cell.alignment)
    filename = file_path  # フルパスを使用
    for offset in range(orig_cols):
        combined_ws.cell(row=start_output_row + offset, column=1, value=filename)
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row >= start_row and merged_range.max_row <= end_row and
            merged_range.min_col >= start_col and merged_range.max_col <= end_col):
            new_start_row = merged_range.min_col - start_col + 1
            new_start_col = merged_range.min_row - start_row + 1
            new_end_row = merged_range.max_col - start_col + 1
            new_end_col = merged_range.max_row - start_row + 1
            final_start_row = start_output_row + new_start_row - 1
            final_start_col = new_start_col + 1  # A列はファイル名用
            final_end_row = start_output_row + new_end_row - 1
            final_end_col = new_end_col + 1
            combined_ws.merge_cells(start_row=final_start_row, start_column=final_start_col,
                                      end_row=final_end_row, end_column=final_end_col)
    return orig_cols

def process_phase3():
    start_time = datetime.now()
    print(f"Phase3 処理開始: {start_time}")
    # Phase2のパターン一覧ファイルをパターン定義として、Phase1のファイル別パターンを元に処理
    pattern_definitions_file = os.path.join(PHASE2_OUTPUT_DIR, f"{MUNICIPALITY_NAME}_パターン一覧_Phase2.xlsx")
    file_list_file = os.path.join(PHASE1_OUTPUT_DIR, f"{MUNICIPALITY_NAME}_ファイル別パターン.xlsx")
    pattern_map = load_pattern_map(pattern_definitions_file, sheet_name="Sheet1")
    if not pattern_map:
        print("パターン定義が空または読み込みに失敗しています。")
        return
    file_list_df = load_file_list(file_list_file, sheet_name="Sheet1")
    if file_list_df is None:
        print("ファイル一覧が空または読み込みに失敗しています。")
        return
    pattern_books = {}
    for idx, row in file_list_df.iterrows():
        municipality = str(row["自治体"]).strip()
        folder_name = str(row["フォルダ名"]).strip()
        file_name = str(row["ファイル名"]).strip()
        pattern_name = str(row["パターン名"]).strip()
        print(f"処理中: {folder_name}\\{file_name}  パターン: {pattern_name}")
        if pattern_name == "なし":
            continue
        if pattern_name not in pattern_map:
            print(f"  → パターン未定義: {pattern_name}")
            continue
        file_path = os.path.join(TARGET_PATH, folder_name, file_name)
        if not os.path.exists(file_path):
            print(f"  → ファイルが見つかりません: {file_path}")
            continue
        if pattern_name not in pattern_books:
            new_wb = Workbook()
            new_ws = new_wb.active
            pattern_books[pattern_name] = {"wb": new_wb, "ws": new_ws, "current_row": 1}
        else:
            new_wb = pattern_books[pattern_name]["wb"]
            new_ws = pattern_books[pattern_name]["ws"]
        current_row = pattern_books[pattern_name]["current_row"]
        try:
            block_rows = process_file_phase3(file_path, new_ws, current_row)
            pattern_books[pattern_name]["current_row"] = current_row + block_rows
        except Exception as e:
            print(f"  → {file_path} の処理に失敗: {e}")
            continue
        time.sleep(1)
    for pattern_name, book_info in pattern_books.items():
        output_file = os.path.join(PHASE3_OUTPUT_DIR, f"{pattern_name}.xlsx")
        try:
            book_info["wb"].save(output_file)
            print(f"パターン【{pattern_name}】の転置データを保存しました: {output_file}")
        except Exception as e:
            print(f"パターン【{pattern_name}】の保存に失敗: {e}")
    end_time = datetime.now()
    print(f"Phase3 処理終了: {end_time}  経過時間: {end_time - start_time}")

import os
from openpyxl import load_workbook, Workbook

def is_effectively_empty(cell_value, col_index=None):
    """
    セルの値がNoneまたは空文字の場合はTrueを返す。
    また、特定の列についてはエラー・プレースホルダー値も空と見なす。
    
    例:
      - Y列（col_index==25）："発送温度帯を選択してください"
      - AN列（col_index==40）："配送方法を選択してください"
    """
    if cell_value is None or cell_value == '':
        return True
    if col_index == 25 and cell_value == "発送温度帯を選択してください":
        return True
    if col_index == 40 and cell_value == "配送方法を選択してください":
        return True
    return False

def combine_rows(row1, row2, col_count):
    """
    2行分のレコードを統合します。
    すべてのセルについて、両行とも非空の場合は文字列として連結、
    片方のみ非空ならその値を、両方空なら空文字とします。
    なお、両者が完全一致している場合は、重複を防いで1つだけ採用します。
    """
    combined = []
    for i in range(col_count):
        val1 = row1[i] if row1[i] is not None else ""
        val2 = row2[i] if row2[i] is not None else ""
        # 両方空なら結果も空
        if not val1 and not val2:
            combined.append("")
        # 両方非空の場合
        elif val1 and val2:
            # 完全一致なら一方のみ
            if str(val1) == str(val2):
                combined.append(val1)
            else:
                # 両方異なる場合は連結（そのまま連結）
                combined.append(str(val1) + str(val2))
        else:
            # 片方だけ非空ならその値
            combined.append(val1 or val2)
    return combined

def process_file_phase4(file_path, output_path):
    # data_only=True でワークブックを読み込み、計算結果（値のみ）を取得
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    print(f"デバッグ: {os.path.basename(file_path)} - Max row: {max_row}, Max col: {max_col}")

    # ① ワークシート全体を値のマトリックスにコピー（値のみ貼り付け）
    data = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        data.append([cell.value for cell in row])
    
    print(f"デバッグ: データ行数: {len(data)}")
    if data:
        print(f"デバッグ: data[0] type: {type(data[0])}, length: {len(data[0]) if isinstance(data[0], list) else 'N/A'}")
        print(f"デバッグ: data[0] content: {data[0][:5] if isinstance(data[0], list) else data[0]}")  # 最初の5要素のみ
        if len(data) > 1:
            print(f"デバッグ: data[1] type: {type(data[1])}, length: {len(data[1]) if isinstance(data[1], list) else 'N/A'}")
            print(f"デバッグ: data[1] content: {data[1][:5] if isinstance(data[1], list) else data[1]}")  # 最初の5要素のみ
    
    # ② 結合セルの解除：各結合範囲について、上位セルの値で全セルを埋める
    for merged_range in ws.merged_cells.ranges:
        r1, r2 = merged_range.min_row, merged_range.max_row
        c1, c2 = merged_range.min_col, merged_range.max_col
        top_left_value = data[r1 - 1][c1 - 1]
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                data[r - 1][c - 1] = top_left_value

    # ③ ヘッダー行とデータ行に分割
    # Phase3の出力構造: 1行目=ファイルパス、2行目=ヘッダー、3行目以降=データ
    if not data:
        print(f"警告: {os.path.basename(file_path)} にデータがありません")
        return
        
    if len(data) < 2:
        print(f"警告: {os.path.basename(file_path)} にヘッダー行がありません")
        return
        
    # Phase3の構造に合わせて、2行目をヘッダーとして使用
    header = data[1]  # 2行目がヘッダー
    # headerがリストであることを確認
    if not isinstance(header, list):
        print(f"警告: {os.path.basename(file_path)} のヘッダーがリストではありません: {type(header)}")
        return
        
    col_count = len(header)
    data_rows = data[2:]  # 3行目以降がデータ行

    # ④ B列（2列目）の値をキーとして、連続する行であれば統合する
    merged_rows = []
    if data_rows:
        current_record = data_rows[0]
        # current_recordがリストであることを確認
        if not isinstance(current_record, list):
            print(f"警告: {os.path.basename(file_path)} の最初のデータ行がリストではありません: {type(current_record)}")
            return
            
        for next_row in data_rows[1:]:
            # next_rowがリストであることを確認
            if not isinstance(next_row, list):
                print(f"警告: {os.path.basename(file_path)} のデータ行がリストではありません: {type(next_row)}")
                continue
                
            # 連続する行かどうかはB列（インデックス1）の値で判断
            if len(next_row) > 1 and len(current_record) > 1 and next_row[1] == current_record[1]:
                # 既に連結されている行同士は、各列のテキストを連結する
                current_record = combine_rows(current_record, next_row, col_count)
            else:
                merged_rows.append(current_record)
                current_record = next_row
        merged_rows.append(current_record)

    # ⑤ 新しいワークブックに出力
    new_wb = Workbook()
    new_ws = new_wb.active
    
    # headerがリストであることを確認してから追加
    if isinstance(header, list):
        new_ws.append(header)
    else:
        print(f"エラー: {os.path.basename(file_path)} のヘッダーがリストではありません")
        return
        
    for record in merged_rows:
        # recordがリストであることを確認してから追加
        if isinstance(record, list):
            new_ws.append(record)
        else:
            print(f"警告: {os.path.basename(file_path)} のレコードがリストではありません: {type(record)}")
            continue
    
    new_wb.save(output_path)
    print(f"正常終了: {os.path.basename(output_path)} を保存しました。")

def process_phase4():
    municipality = os.environ.get("MUNICIPALITY_NAME")
    if not municipality:
        print("MUNICIPALITY_NAME 環境変数が設定されていません。")
        return

    base_dir = os.path.join(
        r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase3\HARV',
        municipality
    )

    if not os.path.exists(base_dir):
        print(f"指定フォルダが存在しません: {base_dir}")
        return

    print(f"MUNICIPALITY_NAME = {municipality}")
    print(f"処理対象フォルダ: {base_dir}")

    files = [f for f in os.listdir(base_dir) if f.startswith("PAT") and f.endswith(".xlsx")]
    if not files:
        print("PATで始まるxlsxファイルが見つかりません。")
        return

    for file in files:
        input_path = os.path.join(base_dir, file)
        output_path = os.path.join(base_dir, file.replace(".xlsx", "_normalized.xlsx"))
        try:
            print(f"処理開始: {file}")
            process_file_phase4(input_path, output_path)
        except Exception as e:
            print(f"エラー: {file} の処理に失敗しました → {e}")
            import traceback
            traceback.print_exc()

import os

from openpyxl import load_workbook, Workbook



# Phase5で不要になった関数群は削除済み



def process_phase5():
    """
    Phase5: Phase4で正規化済みの_normalized.xlsxファイルを直接統合
    """
    municipality = os.environ.get("MUNICIPALITY_NAME")
    if not municipality:
        print("MUNICIPALITY_NAME 環境変数が設定されていません。")
        return

    base_dir = os.path.join(
        r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase3\HARV',
        municipality
    )
    if not os.path.exists(base_dir):
        print(f"指定フォルダが存在しません: {base_dir}")
        return

    print(f"MUNICIPALITY_NAME = {municipality}")
    print(f"処理対象フォルダ: {base_dir}")

    # normalized.xlsx ファイルのみを対象とする（重複正規化ファイルは除外）
    files = [f for f in os.listdir(base_dir) 
             if f.startswith("PAT") and f.endswith("_normalized.xlsx") 
             and not f.endswith("_normalized_normalized.xlsx")]

    if not files:
        print("_normalized.xlsx ファイルが見つかりません。")
        return

    print(f"対象ファイル数: {len(files)}")

    # 全ファイルのデータを直接統合
    all_data = []
    master_headers = None

    for file in files:
        file_path = os.path.join(base_dir, file)
        try:
            # Phase4で正規化済みのファイルを直接読み込み
            df = pd.read_excel(file_path)
            
            if master_headers is None:
                # 最初のファイルのヘッダーをマスターとして採用
                master_headers = list(df.columns)
                print(f"マスターヘッダー設定: {len(master_headers)}列")
            
            # データ行を追加（ヘッダー行は除く）
            for _, row in df.iterrows():
                # 各ファイルの列数をマスターに合わせて調整
                row_data = []
                for header in master_headers:
                    if header in df.columns:
                        row_data.append(row[header])
                    else:
                        row_data.append(None)  # 不足列は空値
                all_data.append(row_data)
            
            print(f"処理完了: {file} ({len(df)}行)")
            
        except Exception as e:
            print(f"エラー: {file} の処理に失敗しました → {e}")

    if not all_data:
        print("統合対象データがありません。")
        return

    # 統合結果をall_collect.xlsxとして保存
    result_df = pd.DataFrame(all_data, columns=master_headers)
    all_collect_path = os.path.join(base_dir, "all_collect.xlsx")
    result_df.to_excel(all_collect_path, index=False)
    
    print(f"all_collect.xlsx 作成完了: {all_collect_path}")
    print(f"統合結果: {len(result_df)}行 × {len(master_headers)}列")

import os
import shutil

def process_phase7():
    # 環境変数からフォルダ名（MUNICIPALITY_NAME）を取得
    municipality = os.environ.get("MUNICIPALITY_NAME")
    if not municipality:
        print("MUNICIPALITY_NAME 環境変数が設定されていません。")
        return

    # Phase3 配下の対象ファイルパスを構築
    source_dir = os.path.join(
        r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase3\HARV',
        municipality
    )
    source_file = os.path.join(source_dir, "all_collect.xlsx")
    
    if not os.path.exists(source_file):
        print(f"ソースファイルが存在しません: {source_file}")
        return

    # 複製先のディレクトリパス
    dest_dir = r"G:\共有ドライブ\★OD\99_商品管理\DATA\Phase4\HARV"
    if not os.path.exists(dest_dir):
        os.makedirs(dest_dir, exist_ok=True)
    
    # 複製後のファイル名を環境変数の値（フォルダ名）に変更
    dest_file = os.path.join(dest_dir, f"{municipality}.xlsx")
    
    try:
        shutil.copy(source_file, dest_file)
        print(f"複製完了: {dest_file}")
    except Exception as e:
        print(f"複製に失敗しました → {e}")

# ===== メイン処理 =====
def main():
    process_phase1(TARGET_PATH, MUNICIPALITY_NAME, PHASE1_OUTPUT_DIR, LOG_FILE_PATH)
    process_phase2(MUNICIPALITY_NAME, PHASE1_OUTPUT_DIR, PHASE2_OUTPUT_DIR)
    process_phase3()
    process_phase4()
    process_phase5()
    process_phase7()

if __name__ == "__main__":
    main()
