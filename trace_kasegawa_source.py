import openpyxl
import os

def trace_kasegawa_in_all_files():
    """
    全ソースファイルで河瀨透の出現位置を追跡
    """
    base_dir = r'DATA\Phase3\HARV\不整合テスト自治体v3'
    
    files = [f for f in os.listdir(base_dir) 
             if f.startswith("PAT") and f.endswith("_normalized.xlsx") 
             and not f.endswith("_normalized_normalized.xlsx")]
    
    print("=== 全ソースファイルでの河瀨透追跡 ===")
    
    kasegawa_sources = {}
    
    for file in sorted(files):
        file_path = os.path.join(base_dir, file)
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            print(f"\n--- {file} ---")
            
            # ヘッダー構造を確認
            header_row1 = [ws.cell(row=1, column=col).value for col in range(1, 30)]
            header_row2 = [ws.cell(row=2, column=col).value for col in range(1, 30)]
            
            # 河瀨透を検索
            kasegawa_positions = []
            for row in range(1, min(50, ws.max_row + 1)):
                for col in range(1, min(30, ws.max_column + 1)):
                    value = ws.cell(row=row, column=col).value
                    if value and '河瀨' in str(value):
                        header1 = header_row1[col-1] if col <= len(header_row1) else None
                        header2 = header_row2[col-1] if col <= len(header_row2) else None
                        kasegawa_positions.append({
                            'row': row,
                            'col': col,
                            'value': value,
                            'header1': header1,
                            'header2': header2
                        })
            
            if kasegawa_positions:
                print(f"  河瀨透発見: {len(kasegawa_positions)}箇所")
                for pos in kasegawa_positions:
                    col_letter = openpyxl.utils.get_column_letter(pos['col'])
                    print(f"    行{pos['row']}, {col_letter}列({pos['col']}): {pos['value']}")
                    print(f"      ヘッダー1: {pos['header1']}")
                    print(f"      ヘッダー2: {pos['header2']}")
                
                kasegawa_sources[file] = kasegawa_positions
            else:
                print("  河瀨透なし")
                
        except Exception as e:
            print(f"  エラー: {e}")
    
    return kasegawa_sources

def identify_problematic_source():
    """
    商品名列に河瀨透を送り込んでいるソースファイルを特定
    """
    # 統合結果ファイルで河瀨透が商品名列にある行を特定
    result_file = r'DATA\Phase3\HARV\不整合テスト自治体v3\perfect_zero_misalignment.xlsx'
    
    wb = openpyxl.load_workbook(result_file, data_only=True)
    ws = wb.active
    
    print("=== 統合結果での河瀨透混入分析 ===")
    
    # 商品名列を特定
    product_name_cols = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and '商品名' in str(header) and 'ご担当者様' not in str(header):
            product_name_cols.append(col)
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"商品名列: {col_letter}列({col}) - {header}")
    
    # 河瀨透が商品名列にある行を分析
    problematic_rows = []
    for row in range(2, ws.max_row + 1):
        for col in product_name_cols:
            value = ws.cell(row=row, column=col).value
            if value and '河瀨' in str(value):
                file_name = ws.cell(row=row, column=1).value
                problematic_rows.append({
                    'row': row,
                    'col': col,
                    'value': value,
                    'source_file': file_name
                })
    
    print(f"\n河瀨透が商品名列に混入している行: {len(problematic_rows)}行")
    for pr in problematic_rows:
        col_letter = openpyxl.utils.get_column_letter(pr['col'])
        print(f"  行{pr['row']}, {col_letter}列: {pr['value']}")
        print(f"    ソースファイル: {os.path.basename(pr['source_file'])}")
    
    return problematic_rows

def create_absolute_perfect_mapping():
    """
    河瀨透問題を完全解決する絶対完璧なマッピングを作成
    """
    # ソースファイルでの河瀨透位置を追跡
    kasegawa_sources = trace_kasegawa_in_all_files()
    
    # 統合結果での問題行を特定
    problematic_rows = identify_problematic_source()
    
    print(f"\n=== 根本原因分析 ===")
    
    # 問題のあるソースファイルを特定
    problem_files = set()
    for pr in problematic_rows:
        source_file = os.path.basename(pr['source_file'])
        if source_file.startswith('PAT'):
            problem_files.add(source_file)
    
    print(f"問題のあるPATファイル: {problem_files}")
    
    # 各問題ファイルの詳細分析
    base_dir = r'DATA\Phase3\HARV\不整合テスト自治体v3'
    absolute_mapping_corrections = {}
    
    for problem_file in problem_files:
        if problem_file in kasegawa_sources:
            print(f"\n--- {problem_file} の詳細分析 ---")
            
            file_path = os.path.join(base_dir, problem_file)
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # ヘッダー行2（サブヘッダー）を詳細確認
            print("サブヘッダー行の詳細:")
            for col in range(1, 20):
                header = ws.cell(row=2, column=col).value
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f"  {col_letter}列({col}): {header}")
            
            # 河瀨透の位置を確認
            kasegawa_positions = kasegawa_sources[problem_file]
            for pos in kasegawa_positions:
                if pos['row'] > 5:  # データ行のみ
                    print(f"\n河瀨透の実際の位置: {openpyxl.utils.get_column_letter(pos['col'])}列({pos['col']})")
                    print(f"  期待されるヘッダー: {pos['header2']}")
                    
                    # この列が本当に担当者名列かチェック
                    if pos['header2'] and 'ご担当者様' in str(pos['header2']):
                        print("  ✅ 正しい位置（ご担当者様列）")
                    elif pos['header2'] and '商品名' in str(pos['header2']):
                        print("  ❌ 間違った位置（商品名列）- マッピング修正が必要")
                        
                        # 正しいご担当者様列を検索
                        correct_col = None
                        for c in range(1, 30):
                            h = ws.cell(row=2, column=c).value
                            if h and 'ご担当者様' in str(h):
                                correct_col = c
                                break
                        
                        if correct_col:
                            print(f"  修正: {openpyxl.utils.get_column_letter(pos['col'])}列 → {openpyxl.utils.get_column_letter(correct_col)}列")
                            
                            # マッピング修正を記録
                            if problem_file not in absolute_mapping_corrections:
                                absolute_mapping_corrections[problem_file] = {}
                            absolute_mapping_corrections[problem_file]['ご担当者様'] = correct_col
    
    print(f"\n=== 絶対完璧マッピング修正 ===")
    for file, corrections in absolute_mapping_corrections.items():
        print(f"{file}: {corrections}")
    
    return absolute_mapping_corrections

if __name__ == "__main__":
    corrections = create_absolute_perfect_mapping()
