import openpyxl
import os

def validate_dynamic_mapping_result():
    """
    動的マッピング結果の検証
    """
    # 統合結果ファイルを開く
    wb = openpyxl.load_workbook('DATA/Phase3/HARV/不整合テスト自治体v3/dynamic_mapping_integration.xlsx', data_only=True)
    ws = wb.active

    print('=== 動的マッピング統合結果の検証 ===')
    print(f'総行数: {ws.max_row}')
    print(f'総列数: {ws.max_column}')

    # ヘッダー行を確認（重要な列のみ）
    print('\n=== 重要ヘッダー確認 ===')
    important_headers = {}
    for col in range(1, min(50, ws.max_column + 1)):  # 最初の50列をチェック
        header = ws.cell(row=1, column=col).value
        if header and any(keyword in str(header) for keyword in ['商品名', 'ご担当者様', '発送元名称', '産地', '返礼品コード']):
            important_headers[col] = header
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f'{col_letter}列({col}): {header}')

    # 河瀨 透 を検索
    print('\n=== 河瀨 透 の検索 ===')
    found_entries = []
    for row in range(2, min(100, ws.max_row + 1)):  # 最初の100行をチェック
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and '河瀨' in str(cell_value):
                header = ws.cell(row=1, column=col).value
                file_name = ws.cell(row=row, column=1).value
                found_entries.append({
                    'row': row, 'col': col, 'header': header,
                    'value': cell_value, 'file': file_name
                })

    print(f'河瀨さんのデータ出現回数: {len(found_entries)}回')
    for entry in found_entries:
        col_letter = openpyxl.utils.get_column_letter(entry['col'])
        print(f'  行{entry["row"]}, {col_letter}列({entry["header"]}): {entry["value"]} [ファイル: {entry["file"]}]')

    # PAT0001のデータサンプルを確認
    print('\n=== PAT0001 データサンプル ===')
    pat0001_rows = []
    for row in range(2, min(20, ws.max_row + 1)):
        file_name = ws.cell(row=row, column=1).value
        if file_name and 'PAT0001' in str(file_name):
            pat0001_rows.append(row)

    if pat0001_rows:
        sample_row = pat0001_rows[0]
        print(f'PAT0001 サンプル行{sample_row}:')
        for col in important_headers:
            value = ws.cell(row=sample_row, column=col).value
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f'  {col_letter}列 {important_headers[col]}: {value}')
    
    # ヘッダーの重複確認
    print('\n=== ヘッダー重複確認 ===')
    header_counts = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            key = str(header).strip()
            if key in header_counts:
                header_counts[key] += 1
            else:
                header_counts[key] = 1
    
    duplicates = {k: v for k, v in header_counts.items() if v > 1}
    if duplicates:
        print(f'重複ヘッダー数: {len(duplicates)}')
        for header, count in duplicates.items():
            print(f'  {header}: {count}回')
    else:
        print('重複ヘッダーなし ✅')
    
    return {
        'total_rows': ws.max_row,
        'total_columns': ws.max_column,
        'important_headers': important_headers,
        'kasegawa_entries': found_entries,
        'pat0001_sample_row': pat0001_rows[0] if pat0001_rows else None,
        'duplicate_headers': duplicates
    }

if __name__ == "__main__":
    result = validate_dynamic_mapping_result()
