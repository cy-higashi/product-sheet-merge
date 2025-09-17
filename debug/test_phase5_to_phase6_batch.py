import os

from openpyxl import load_workbook, Workbook



def update_master_headers(master_headers, source_headers):

  """

  source_headers 内の各項目について、master_headers に

  完全一致するものがなければ追加する。

  （大文字小文字も含めた完全一致判定を行います）

  """

  for header in source_headers:

    if header not in master_headers:

      master_headers.append(header)

  return master_headers



def process_file(file_path, master_headers):

  """

  １ファイルを処理します。

 

  ・対象ファイルは、既に normalized 状態であるため、結合解除等の処理は不要。

  ・ワークシート内でB列に「項目」と記載された行をヘッダー行とみなし、

   その行のC列以降のセルをソースヘッダーとして取得します。

  ・master_headers（マスター側のヘッダーリスト）は、ソースヘッダーと完全一致するものはそのまま、

   異なる文字列（１文字でも違えば）は新規追加します。

  ・ヘッダー行の下のデータ行について、出力用の1行データを作成します。

   ※ 出力行のA列にはファイル名、B列には元ファイルのB列の値をセットし、

    以降、master_headers の各項目に対応するデータを配置します。

  """

  wb = load_workbook(file_path, data_only=True)

  ws = wb.active

  max_row = ws.max_row

  max_col = ws.max_column



  # ワークシート全体の値を2次元リストにコピー（前処理済みのため、直接 cell.value を利用）

  data = []

  for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):

    data.append([cell.value for cell in row])

 

  # ヘッダー行（B列が「項目」）を検索

  header_row_index = None

  for i, row in enumerate(data):

    if len(row) >= 2 and row[1] == "項目":

      header_row_index = i

      break

  if header_row_index is None:

    print(f"ファイル内に「項目」行が見つかりません: {os.path.basename(file_path)} → このファイルはスキップします。")

    return [], master_headers



  # ソースヘッダーは、ヘッダー行のC列以降（インデックス2以降）のセル値（None 以外）とする

  source_headers = [cell for cell in data[header_row_index][2:] if cell is not None]

  master_headers = update_master_headers(master_headers, source_headers)



  # ソースファイル側のヘッダーと、その列番号（C列以降）との対応を作成

  # ※ ヘッダー行は data[header_row_index]、C列以降は列番号 2～max_col（0-index: 2～）

  source_header_to_index = {}

  for j, header in enumerate(data[header_row_index][2:], start=2):

    if header is not None:

      source_header_to_index[header] = j



  # ヘッダー行の下（header_row_index+1 以降）がデータ行

  output_rows = []

  for row in data[header_row_index+1:]:

    new_row = []

    # A列: 既存のファイル名を保持（元のファイル名）

    # Phase3で設定された元のファイル名を保持し、PAT0001_normalized.xlsx などで上書きしない

    existing_file_name = row[0] if len(row) > 0 and row[0] is not None else os.path.basename(file_path)

    new_row.append(existing_file_name)

    # B列: 元ファイルのB列の値（例えば、項目の名称など）

    value_B = row[1] if len(row) > 1 else None

    new_row.append(value_B)

    # C列以降: master_headers に従い、該当するデータがあれば配置、なければ None

    for header in master_headers:

      if header in source_header_to_index:

        idx = source_header_to_index[header]

        new_row.append(row[idx] if idx < len(row) else None)

      else:

        new_row.append(None)

    output_rows.append(new_row)

  return output_rows, master_headers



def main():

  municipality = os.environ.get("MUNICIPALITY_NAME")

  if not municipality:

    print("MUNICIPALITY_NAME 環境変数が設定されていません。")

    return



  base_dir = os.path.join(

    r'G:\共有ドライブ\★OD_管理者\データマネジメント部\DataOps\オペレーション\商品管理\2025-04\Phase3\HARV',

    municipality

  )

  if not os.path.exists(base_dir):

    print(f"指定フォルダが存在しません: {base_dir}")

    return



  print(f"MUNICIPALITY_NAME = {municipality}")

  print(f"処理対象フォルダ: {base_dir}")



  # normalized.xlsx ファイルのみを対象とする

  files = [f for f in os.listdir(base_dir) if f.startswith("PAT") and f.endswith("_normalized.xlsx")]

  if not files:

    print("_normalized.xlsx ファイルが見つかりません。")

    return



  # マスター集約用ワークブック作成（all_collect.xlsx）

  master_wb = Workbook()

  master_ws = master_wb.active

  master_ws.title = "all_collect"

  # 固定ヘッダー：A列＝ファイル名、B列＝項目

  master_ws.cell(row=1, column=1, value="ファイル名")

  master_ws.cell(row=1, column=2, value="項目")

  # master_headers：C列以降の項目（Union したすべてのソースヘッダー）

  master_headers = []

  master_data_rows = [] # 各ファイルから抽出したデータ行を集約



  for file in files:

    file_path = os.path.join(base_dir, file)

    try:

      rows, master_headers = process_file(file_path, master_headers)

      master_data_rows.extend(rows)

      print(f"処理完了: {file}")

    except Exception as e:

      print(f"エラー: {file} の処理に失敗しました → {e}")



  # マスターシートの1行目（固定ヘッダーに続くC列以降）に master_headers を設定

  for i, header in enumerate(master_headers, start=3):

    master_ws.cell(row=1, column=i, value=header)



  # 集約したデータ行をマスターシートに書き出す（2行目以降）

  current_row = 2

  for row in master_data_rows:

    for j, value in enumerate(row, start=1):

      master_ws.cell(row=current_row, column=j, value=value)

    current_row += 1



  all_collect_path = os.path.join(base_dir, "all_collect.xlsx")

  master_wb.save(all_collect_path)

  print(f"all_collect.xlsx 作成完了: {all_collect_path}")



if __name__ == "__main__":

  main()