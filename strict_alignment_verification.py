import openpyxl
import os
import json
from collections import defaultdict
import re

def analyze_source_file_headers(file_path):
    """
    ソースファイルの実際のヘッダー構造を詳細分析
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        max_row = ws.max_row
        max_col = ws.max_column

        # ワークシート全体の値を2次元リストにコピー
        data = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            data.append([cell.value for cell in row])

        # 「項目」行を検索
        header_row_index = None
        for i, row in enumerate(data):
            if len(row) >= 2 and row[1] == "項目":
                header_row_index = i
                break
        
        if header_row_index is None:
            return None
        
        # ヘッダー行とサブヘッダー行を取得
        header_row = data[header_row_index] if header_row_index < len(data) else []
        sub_header_row = data[header_row_index + 1] if header_row_index + 1 < len(data) else []
        
        # データ行のサンプルを取得
        sample_data_rows = []
        for row_idx in range(header_row_index + 2, min(header_row_index + 7, len(data))):
            if row_idx < len(data):
                sample_data_rows.append(data[row_idx])
        
        return {
            'file_name': os.path.basename(file_path),
            'header_row_index': header_row_index,
            'header_row': header_row,
            'sub_header_row': sub_header_row,
            'sample_data_rows': sample_data_rows,
            'max_col': max_col
        }
        
    except Exception as e:
        print(f"  ❌ ファイル分析エラー: {file_path} → {e}")
        return None

def detect_semantic_misalignment(source_analysis, integrated_result):
    """
    セマンティック（意味的）な列ズレを検出
    """
    misalignments = []
    
    # 重要なデータ型パターンを定義
    patterns = {
        'person_name': r'^[一-龯]{2,4}\s*[一-龯]{1,3}$',  # 日本人名パターン
        'product_name': r'(肉|米|野菜|果物|魚|酒|茶|菓子|パン|麺|調味料)',  # 商品名キーワード
        'phone_number': r'^\d{2,4}-\d{2,4}-\d{4}$',  # 電話番号パターン
        'address': r'(市|町|村|区|県|都|府|道)',  # 住所キーワード
        'code': r'^[A-Z0-9]{3,10}$',  # コードパターン
        'amount': r'^\d+g$|^\d+ml$|^\d+個$',  # 内容量パターン
    }
    
    wb = openpyxl.load_workbook(integrated_result, data_only=True)
    ws = wb.active
    
    # ヘッダー情報を取得
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers[col] = str(header)
    
    # 重要な列の期待データ型を定義
    expected_data_types = {
        'ご担当者様': 'person_name',
        '商品名': 'product_name',
        '事業者様TEL': 'phone_number',
        '発送元TEL': 'phone_number',
        '発送元住所': 'address',
        '返礼品コード': 'code',
        '内容量': 'amount'
    }
    
    # 各重要列のデータ型を検証
    for col, header in headers.items():
        # ヘッダー名が期待データ型リストにあるかチェック
        expected_type = None
        for expected_header, data_type in expected_data_types.items():
            if expected_header in header:
                expected_type = data_type
                break
        
        if expected_type:
            # この列のデータサンプルを取得
            sample_values = []
            for row in range(2, min(50, ws.max_row + 1)):
                value = ws.cell(row=row, column=col).value
                if value and str(value).strip():
                    sample_values.append(str(value).strip())
            
            # データ型パターンマッチング
            pattern = patterns[expected_type]
            mismatched_count = 0
            total_count = len(sample_values)
            
            if total_count > 0:
                for value in sample_values:
                    if not re.match(pattern, value):
                        mismatched_count += 1
                
                mismatch_rate = (mismatched_count / total_count) * 100
                
                if mismatch_rate > 10:  # 10%以上のミスマッチで警告
                    misalignments.append({
                        'column': col,
                        'header': header,
                        'expected_type': expected_type,
                        'mismatch_rate': mismatch_rate,
                        'sample_mismatches': [v for v in sample_values[:5] if not re.match(pattern, v)]
                    })
    
    return misalignments

def cross_validate_with_source_files(base_dir, integrated_result):
    """
    ソースファイルと統合結果の交差検証
    """
    # 対象ファイルを取得
    files = [f for f in os.listdir(base_dir) 
             if f.startswith("PAT") and f.endswith("_normalized.xlsx") 
             and not f.endswith("_normalized_normalized.xlsx")]
    
    print(f"=== 厳密な列ズレ検証 ===")
    print(f"対象ファイル: {len(files)}個")
    
    # 各ソースファイルの構造を分析
    source_analyses = {}
    for file in sorted(files):
        file_path = os.path.join(base_dir, file)
        analysis = analyze_source_file_headers(file_path)
        if analysis:
            source_analyses[file] = analysis
            print(f"✅ {file}: ヘッダー行{analysis['header_row_index'] + 1}, 最大列{analysis['max_col']}")
    
    # セマンティック列ズレを検出
    print(f"\n=== セマンティック列ズレ検出 ===")
    misalignments = detect_semantic_misalignment(source_analyses, integrated_result)
    
    if misalignments:
        print(f"⚠️ 検出された列ズレ: {len(misalignments)}件")
        for mis in misalignments:
            print(f"  列{mis['column']} ({mis['header']}): {mis['mismatch_rate']:.1f}%のミスマッチ")
            print(f"    期待データ型: {mis['expected_type']}")
            print(f"    サンプルミスマッチ: {mis['sample_mismatches'][:3]}")
    else:
        print("✅ セマンティック列ズレなし (0%)")
    
    # 人名の位置確認（河瀨透問題の詳細分析）
    print(f"\n=== 人名データの位置確認 ===")
    wb = openpyxl.load_workbook(integrated_result, data_only=True)
    ws = wb.active
    
    person_name_columns = []
    product_name_columns = []
    
    # 人名と商品名の列を特定
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            if 'ご担当者様' in str(header) or '担当者' in str(header):
                person_name_columns.append(col)
            elif '商品名' in str(header):
                product_name_columns.append(col)
    
    print(f"人名列: {[f'{openpyxl.utils.get_column_letter(c)}列' for c in person_name_columns]}")
    print(f"商品名列: {[f'{openpyxl.utils.get_column_letter(c)}列' for c in product_name_columns]}")
    
    # 河瀨透の出現位置を詳細分析
    kasegawa_positions = []
    for row in range(2, min(100, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if value and '河瀨' in str(value):
                header = ws.cell(row=1, column=col).value
                file_name = ws.cell(row=row, column=1).value
                kasegawa_positions.append({
                    'row': row, 'col': col, 'header': header,
                    'value': value, 'file': file_name,
                    'is_person_column': col in person_name_columns,
                    'is_product_column': col in product_name_columns
                })
    
    # 列ズレ率を計算
    total_kasegawa = len(kasegawa_positions)
    misplaced_kasegawa = len([k for k in kasegawa_positions if k['is_product_column']])
    
    if total_kasegawa > 0:
        misplacement_rate = (misplaced_kasegawa / total_kasegawa) * 100
        print(f"\n河瀨透データ分析:")
        print(f"  総出現回数: {total_kasegawa}")
        print(f"  正しい位置(人名列): {total_kasegawa - misplaced_kasegawa}回")
        print(f"  間違った位置(商品名列): {misplaced_kasegawa}回")
        print(f"  列ズレ率: {misplacement_rate:.2f}%")
        
        if misplaced_kasegawa > 0:
            print(f"\n  間違った位置の詳細:")
            for k in kasegawa_positions:
                if k['is_product_column']:
                    col_letter = openpyxl.utils.get_column_letter(k['col'])
                    print(f"    行{k['row']}, {col_letter}列({k['header']}): {k['value']} [ファイル: {os.path.basename(k['file'])}]")
    
    return {
        'total_files': len(files),
        'semantic_misalignments': misalignments,
        'kasegawa_analysis': {
            'total': total_kasegawa,
            'misplaced': misplaced_kasegawa,
            'misplacement_rate': misplacement_rate if total_kasegawa > 0 else 0
        }
    }

def strict_alignment_verification(municipality_name="不整合テスト自治体v3"):
    """
    列ズレ0%達成の厳密な検証
    """
    base_dir = os.path.join(
        r'DATA\Phase3\HARV',
        municipality_name
    )
    
    if not os.path.exists(base_dir):
        print(f"指定フォルダが存在しません: {base_dir}")
        return None
    
    integrated_result = os.path.join(base_dir, "dynamic_mapping_integration.xlsx")
    
    if not os.path.exists(integrated_result):
        print(f"統合結果ファイルが存在しません: {integrated_result}")
        return None
    
    # 交差検証を実行
    result = cross_validate_with_source_files(base_dir, integrated_result)
    
    # 最終判定
    print(f"\n=== 最終列ズレ判定 ===")
    
    semantic_errors = len(result['semantic_misalignments'])
    kasegawa_misplacement_rate = result['kasegawa_analysis']['misplacement_rate']
    
    total_error_rate = max(kasegawa_misplacement_rate, 
                          (semantic_errors / result['total_files']) * 100 if result['total_files'] > 0 else 0)
    
    if total_error_rate == 0:
        print("🎉 列ズレ 0.00% - 完璧な整列達成！")
        verdict = "PERFECT"
    elif total_error_rate < 1:
        print(f"⚠️ 列ズレ {total_error_rate:.2f}% - 軽微な問題あり")
        verdict = "MINOR_ISSUES"
    else:
        print(f"❌ 列ズレ {total_error_rate:.2f}% - 要修正")
        verdict = "NEEDS_FIX"
    
    return {
        'verdict': verdict,
        'total_error_rate': total_error_rate,
        'details': result
    }

if __name__ == "__main__":
    result = strict_alignment_verification()
