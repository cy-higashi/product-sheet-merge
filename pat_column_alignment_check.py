import os
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict
import re

def analyze_column_meanings(municipality_name="不整合テスト自治体v3"):
    """
    PAT間での列の意味的ズレを検出・分析するスクリプト
    """
    base_dir = os.path.join(
        r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase3\HARV',
        municipality_name
    )
    
    if not os.path.exists(base_dir):
        print(f"指定フォルダが存在しません: {base_dir}")
        return
    
    # 対象ファイルを取得（重複正規化ファイルは除外）
    files = [f for f in os.listdir(base_dir) 
             if f.startswith("PAT") and f.endswith("_normalized.xlsx") 
             and not f.endswith("_normalized_normalized.xlsx")]
    
    if not files:
        print("対象ファイルが見つかりません。")
        return
    
    print(f"分析対象ファイル: {len(files)}個")
    print("=" * 100)
    
    # 各PATファイルのヘッダー情報を収集
    pat_headers = {}
    pat_sample_data = {}
    
    for file in sorted(files):
        file_path = os.path.join(base_dir, file)
        print(f"\n=== {file} の分析 ===")
        
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            max_row = ws.max_row
            max_col = ws.max_column

            # ワークシート全体の値を2次元リストにコピー
            data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                data.append([cell.value for cell in row])

            # ヘッダー行（B列が「項目」）を検索
            header_row_index = None
            for i, row in enumerate(data):
                if len(row) >= 2 and row[1] == "項目":
                    header_row_index = i
                    break

            if header_row_index is None:
                print(f"  ⚠️ 「項目」行が見つかりません")
                continue

            # ヘッダーを抽出
            headers = [cell for cell in data[header_row_index][2:] if cell is not None]
            print(f"  📋 ヘッダー数: {len(headers)}")
            
            # サンプルデータを抽出（データ行の最初の数行）
            sample_data = []
            data_start = header_row_index + 1
            for row_idx in range(data_start, min(data_start + 5, len(data))):
                if row_idx < len(data):
                    row_data = data[row_idx][2:]  # C列以降
                    sample_data.append(row_data)
            
            pat_headers[file] = headers
            pat_sample_data[file] = sample_data
            
            # ヘッダーの一部を表示
            print(f"  📝 ヘッダー例（最初の10個）:")
            for i, header in enumerate(headers[:10]):
                print(f"    [{i+3}列目] {header}")
            
        except Exception as e:
            print(f"  ❌ エラー: {e}")
    
    print("\n" + "=" * 100)
    print("=== PAT間でのヘッダー位置比較 ===")
    
    # 共通ヘッダーを特定
    all_headers = set()
    for headers in pat_headers.values():
        all_headers.update(headers)
    
    common_headers = []
    for header in all_headers:
        count = sum(1 for headers in pat_headers.values() if header in headers)
        if count >= 2:  # 2つ以上のPATで使用されているヘッダー
            common_headers.append((header, count))
    
    # 使用頻度順でソート
    common_headers.sort(key=lambda x: x[1], reverse=True)
    
    print(f"\n📊 複数PATで使用されているヘッダー: {len(common_headers)}個")
    
    # 位置ズレの検出
    alignment_issues = []
    
    for header, usage_count in common_headers[:20]:  # 上位20個をチェック
        positions = {}
        for file, headers in pat_headers.items():
            if header in headers:
                pos = headers.index(header)
                positions[file] = pos
        
        if len(set(positions.values())) > 1:  # 位置が異なる場合
            alignment_issues.append((header, positions))
    
    print(f"\n🚨 位置ズレが検出されたヘッダー: {len(alignment_issues)}個")
    
    # 詳細分析
    critical_issues = []
    
    for header, positions in alignment_issues:
        print(f"\n--- ヘッダー: 「{header}」 ---")
        
        # 各PATでの位置とサンプルデータを表示
        sample_values = {}
        for file, pos in positions.items():
            print(f"  {file}: {pos+3}列目（C列を3列目とする）")
            
            # そのヘッダー位置のサンプルデータを取得
            if file in pat_sample_data:
                values = []
                for sample_row in pat_sample_data[file]:
                    if pos < len(sample_row) and sample_row[pos] is not None:
                        val = str(sample_row[pos]).strip()
                        if val and val != "":
                            values.append(val)
                
                sample_values[file] = values[:3]  # 最初の3個
                print(f"    サンプル: {values[:3]}")
        
        # データの内容を比較して意味的な違いを検出
        is_critical = detect_semantic_mismatch(header, sample_values)
        if is_critical:
            critical_issues.append((header, positions, sample_values))
    
    print("\n" + "=" * 100)
    print("=== 🔥 重大な列ズレ問題 ===")
    
    if critical_issues:
        for i, (header, positions, sample_values) in enumerate(critical_issues, 1):
            print(f"\n{i}. 【{header}】")
            print("  異なる意味のデータが同じヘッダーに統合されています：")
            for file, values in sample_values.items():
                pos = positions[file]
                print(f"    {file} ({pos+3}列目): {values}")
            
            print("  💡 推奨対応:")
            print("    - 各PATファイルでのヘッダー名を統一")
            print("    - または、ヘッダー位置を揃える")
    else:
        print("重大な意味的ズレは検出されませんでした。")
    
    print(f"\n📈 分析サマリー:")
    print(f"  - 分析対象PAT数: {len(pat_headers)}")
    print(f"  - 共通ヘッダー数: {len(common_headers)}")
    print(f"  - 位置ズレヘッダー数: {len(alignment_issues)}")
    print(f"  - 重大問題数: {len(critical_issues)}")
    
    # 詳細レポートをCSVで出力
    save_detailed_report(base_dir, pat_headers, alignment_issues, sample_values, municipality_name)

def detect_semantic_mismatch(header, sample_values):
    """
    サンプルデータから意味的な不一致を検出
    """
    if len(sample_values) < 2:
        return False
    
    # データタイプの分析
    types_analysis = {}
    for file, values in sample_values.items():
        analysis = analyze_data_type(values)
        types_analysis[file] = analysis
    
    # 異なるタイプが混在している場合は重大
    unique_types = set()
    for analysis in types_analysis.values():
        unique_types.add(analysis['primary_type'])
    
    if len(unique_types) > 1:
        print(f"    🚨 データタイプの不一致検出: {unique_types}")
        return True
    
    # 特定の人名パターンと商品名パターンの混在チェック
    has_person_name = False
    has_product_name = False
    
    for values in sample_values.values():
        for val in values:
            if is_likely_person_name(val):
                has_person_name = True
            if is_likely_product_name(val):
                has_product_name = True
    
    if has_person_name and has_product_name:
        print(f"    🚨 人名と商品名の混在検出")
        return True
    
    return False

def analyze_data_type(values):
    """
    データの種類を分析
    """
    if not values:
        return {'primary_type': 'empty'}
    
    numeric_count = 0
    date_count = 0
    text_count = 0
    
    for val in values:
        val_str = str(val).strip()
        if not val_str:
            continue
            
        # 数値チェック
        if re.match(r'^\d+$', val_str) or re.match(r'^\d+\.\d+$', val_str):
            numeric_count += 1
        # 日付チェック
        elif re.match(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', val_str):
            date_count += 1
        # その他はテキスト
        else:
            text_count += 1
    
    total = numeric_count + date_count + text_count
    if total == 0:
        return {'primary_type': 'empty'}
    
    if numeric_count / total > 0.6:
        return {'primary_type': 'numeric', 'confidence': numeric_count / total}
    elif date_count / total > 0.6:
        return {'primary_type': 'date', 'confidence': date_count / total}
    else:
        return {'primary_type': 'text', 'confidence': text_count / total}

def is_likely_person_name(value):
    """
    人名らしいかどうかを判定
    """
    if not value:
        return False
    
    val = str(value).strip()
    
    # 日本人名のパターン
    if re.match(r'^[ぁ-んァ-ヶ一-龯]{2,6}[\s　]*[ぁ-んァ-ヶ一-龯]{1,6}$', val):
        return True
    
    # 「河瀨　透」のような具体例
    person_patterns = ['河瀨', '透', '田中', '佐藤', '山田', '高橋', '松本']
    return any(pattern in val for pattern in person_patterns)

def is_likely_product_name(value):
    """
    商品名らしいかどうかを判定
    """
    if not value:
        return False
    
    val = str(value).strip()
    
    # 商品名のパターン
    product_keywords = ['kg', 'g', 'ml', 'L', '個', 'セット', '詰合せ', '牛肉', '米', 'みかん', 'いちご']
    return any(keyword in val for keyword in product_keywords)

def save_detailed_report(base_dir, pat_headers, alignment_issues, sample_values, municipality_name):
    """
    詳細レポートをCSVファイルで保存
    """
    import csv
    
    report_path = os.path.join(base_dir, f"{municipality_name}_column_alignment_report.csv")
    
    with open(report_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # ヘッダー行
        writer.writerow(['ヘッダー名', 'PAT1ファイル', 'PAT1位置', 'PAT1サンプル', 
                        'PAT2ファイル', 'PAT2位置', 'PAT2サンプル', '問題レベル'])
        
        # データ行
        for header, positions in alignment_issues:
            files = list(positions.keys())
            if len(files) >= 2:
                file1, file2 = files[0], files[1]
                pos1, pos2 = positions[file1], positions[file2]
                
                # サンプルデータ取得
                sample1 = sample_values.get(file1, []) if 'sample_values' in locals() else []
                sample2 = sample_values.get(file2, []) if 'sample_values' in locals() else []
                
                sample1_str = ' | '.join(map(str, sample1[:2]))
                sample2_str = ' | '.join(map(str, sample2[:2]))
                
                # 問題レベル判定
                level = "高" if detect_semantic_mismatch(header, {file1: sample1, file2: sample2}) else "中"
                
                writer.writerow([header, file1, pos1+3, sample1_str, 
                                file2, pos2+3, sample2_str, level])
    
    print(f"\n📄 詳細レポートを保存しました: {report_path}")

if __name__ == "__main__":
    analyze_column_meanings()
