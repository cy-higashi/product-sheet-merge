#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
カラムズレ問題再現・検証スクリプト

実際のmerge.pyの処理を模擬して、カラムズレが発生する条件を特定する
"""

import os
import pandas as pd
from openpyxl import load_workbook
import re
from datetime import datetime

class ColumnMisalignmentReproducer:
    def __init__(self):
        self.debug_info = []
        self.municipality = "2025-09-09_福岡県北九州市"
    
    def log_debug(self, message, level="INFO"):
        """デバッグ情報をログ"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] [{level}] {message}"
        self.debug_info.append(log_entry)
        print(log_entry)
    
    def simulate_find_header_base_position(self, data):
        """
        merge.pyのfind_header_base_position関数を模擬
        問題の原因を特定するため詳細ログを追加
        """
        self.log_debug("=== ヘッダー基準位置検出開始 ===")
        
        header_row_index = None
        header_col_index = None
        
        # 「項目」とある行を検索
        item_candidates = []
        for i, row in enumerate(data):
            if len(row) >= 2:
                cell_value = row[1] if row[1] is not None else ""
                if str(cell_value) == "項目":
                    item_candidates.append(i)
                    self.log_debug(f"「項目」発見: 行{i+1}, 値='{cell_value}'")
        
        if item_candidates:
            header_row_index = item_candidates[0]  # 最初に見つかったものを使用
            self.log_debug(f"選択された項目行: {header_row_index + 1}")
            
            if len(item_candidates) > 1:
                self.log_debug(f"警告: 複数の「項目」が見つかりました: {[i+1 for i in item_candidates]}", "WARNING")
        else:
            self.log_debug("「項目」が見つかりませんでした", "ERROR")
            return None, None
        
        # 「返礼品コード」がある列を検索（ヘッダー行内で）
        code_candidates = []
        if header_row_index is not None:
            for j, cell in enumerate(data[header_row_index]):
                if cell and "返礼品コード" in str(cell):
                    code_candidates.append(j)
                    self.log_debug(f"「返礼品コード」発見: 列{j+1}, 値='{cell}'")
        
        if code_candidates:
            header_col_index = code_candidates[0]  # 最初に見つかったものを使用
            self.log_debug(f"選択された返礼品コード列: {header_col_index + 1}")
            
            if len(code_candidates) > 1:
                self.log_debug(f"警告: 複数の「返礼品コード」が見つかりました: {[j+1 for j in code_candidates]}", "WARNING")
        else:
            self.log_debug("「返礼品コード」が見つかりませんでした", "ERROR")
            return header_row_index, None
        
        self.log_debug(f"最終結果: 行{header_row_index + 1}, 列{header_col_index + 1}")
        return header_row_index, header_col_index
    
    def simulate_extract_dynamic_headers(self, data, header_row_index, header_col_start):
        """
        merge.pyのextract_dynamic_headers関数を模擬
        """
        self.log_debug("=== 動的ヘッダー抽出開始 ===")
        
        if header_row_index is None or header_col_start is None:
            self.log_debug("ヘッダー基準位置が不正です", "ERROR")
            return []
        
        main_row = data[header_row_index]
        sub_row = data[header_row_index + 1] if header_row_index + 1 < len(data) else []
        
        self.log_debug(f"メインヘッダー行: {len(main_row)}列")
        self.log_debug(f"サブヘッダー行: {len(sub_row)}列")
        
        dynamic_headers = []
        max_col = max(len(main_row), len(sub_row))
        
        for col_idx in range(header_col_start, max_col):
            main_header = main_row[col_idx] if col_idx < len(main_row) else None
            sub_header = sub_row[col_idx] if col_idx < len(sub_row) else None
            
            # 階層的ヘッダー名を生成
            header_name = self.simulate_create_hierarchical_header(
                main_header, sub_header, col_idx, header_col_start)
            
            dynamic_headers.append({
                'column_index': col_idx + 1,  # 1-based
                'header_name': header_name,
                'main_header': main_header,
                'sub_header': sub_header
            })
            
            self.log_debug(f"列{col_idx + 1}: '{main_header}' + '{sub_header}' → '{header_name}'")
        
        self.log_debug(f"動的ヘッダー数: {len(dynamic_headers)}")
        return dynamic_headers
    
    def simulate_create_hierarchical_header(self, main_header, sub_header, col_index, header_col_start):
        """
        merge.pyのcreate_hierarchical_header関数を模擬
        """
        main_clean = self.normalize_header_text(main_header) if main_header else ""
        sub_clean = self.normalize_header_text(sub_header) if sub_header else ""
        
        # 階層構造の処理
        if main_clean and sub_clean and sub_clean not in ["必須", "任意"]:
            result = f"{main_clean}:{sub_clean}"
        elif main_clean:
            result = main_clean
        elif sub_clean and sub_clean not in ["必須", "任意"]:
            result = sub_clean
        else:
            # 空の場合は列位置で識別
            col_relative = col_index - header_col_start + 1
            result = f"空列{col_relative}"
        
        # 問題のあるパターンを検出
        if main_clean and sub_clean:
            if "商品名" in main_clean and any(pattern in sub_clean for pattern in ["担当者", "氏名", "名前"]):
                self.log_debug(f"警告: 商品名列に人名サブヘッダー検出 - '{main_clean}:{sub_clean}'", "WARNING")
            elif "内容量" in main_clean and any(pattern in sub_clean for pattern in ["会社", "企業", "法人"]):
                self.log_debug(f"警告: 内容量列に会社名サブヘッダー検出 - '{main_clean}:{sub_clean}'", "WARNING")
        
        return result
    
    def normalize_header_text(self, text):
        """
        merge.pyのnormalize_header_text関数を模擬
        """
        if not text:
            return ""
        text = str(text).strip()
        # 改行、連続空白を除去
        text = re.sub(r'\s+', '', text)
        return text
    
    def test_problematic_file(self, file_path):
        """
        問題のあるファイルでテスト実行
        """
        self.log_debug(f"=== ファイルテスト開始: {os.path.basename(file_path)} ===")
        
        if not os.path.exists(file_path):
            self.log_debug(f"ファイルが存在しません: {file_path}", "ERROR")
            return None
        
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            max_row = ws.max_row
            max_col = ws.max_column
            
            self.log_debug(f"ファイル情報: {max_row}行 × {max_col}列")
            
            # ワークシート全体の値を2次元リストにコピー
            data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                data.append([cell.value for cell in row])
            
            # 結合セルの解除
            merged_ranges = list(ws.merged_cells.ranges)
            self.log_debug(f"結合セル数: {len(merged_ranges)}")
            
            for merged_range in merged_ranges:
                r1, r2 = merged_range.min_row, merged_range.max_row
                c1, c2 = merged_range.min_col, merged_range.max_col
                top_left_value = data[r1 - 1][c1 - 1]
                for r in range(r1, r2 + 1):
                    for c in range(c1, c2 + 1):
                        data[r - 1][c - 1] = top_left_value
            
            # ヘッダー基準位置の検出
            header_row_index, header_col_start = self.simulate_find_header_base_position(data)
            
            if header_row_index is None or header_col_start is None:
                self.log_debug("ヘッダー基準位置の検出に失敗", "ERROR")
                return None
            
            # 動的ヘッダー構造の解析
            dynamic_headers = self.simulate_extract_dynamic_headers(data, header_row_index, header_col_start)
            
            # データ行の分析
            data_rows = data[header_row_index + 2:]  # ヘッダー行+サブヘッダー行の次から
            self.log_debug(f"データ行数: {len(data_rows)}")
            
            # 疑わしいデータパターンをチェック
            self.check_suspicious_patterns(dynamic_headers, data_rows)
            
            result = {
                'file_path': file_path,
                'header_row_index': header_row_index,
                'header_col_start': header_col_start,
                'dynamic_headers': dynamic_headers,
                'data_row_count': len(data_rows),
                'sample_data': data_rows[:5] if data_rows else []
            }
            
            return result
            
        except Exception as e:
            self.log_debug(f"ファイル処理エラー: {e}", "ERROR")
            return None
    
    def check_suspicious_patterns(self, dynamic_headers, data_rows):
        """
        疑わしいデータパターンをチェック
        """
        self.log_debug("=== 疑わしいパターンチェック ===")
        
        # ヘッダー名から列インデックスのマッピング
        header_to_col = {}
        for header_info in dynamic_headers:
            header_name = header_info['header_name']
            col_index = header_info['column_index'] - 1  # 0-based
            header_to_col[header_name] = col_index
        
        suspicious_count = 0
        
        for i, row in enumerate(data_rows[:20]):  # 最初の20行をチェック
            if len(row) <= 2:
                continue
                
            for header_name, col_index in header_to_col.items():
                if col_index >= len(row):
                    continue
                    
                cell_value = row[col_index]
                if not cell_value:
                    continue
                    
                cell_str = str(cell_value)
                
                # 商品名列に人名パターン
                if "商品名" in header_name:
                    name_patterns = ["河瀨", "担当者", "氏名", "様", "さん"]
                    for pattern in name_patterns:
                        if pattern in cell_str:
                            self.log_debug(f"疑わしいパターン: 行{i+3}, 列{header_name} = '{cell_str}' (パターン: {pattern})", "WARNING")
                            suspicious_count += 1
                
                # 内容量列に会社名パターン
                elif "内容量" in header_name or "容量" in header_name:
                    company_patterns = ["株式会社", "有限会社", "合同会社"]
                    for pattern in company_patterns:
                        if pattern in cell_str:
                            self.log_debug(f"疑わしいパターン: 行{i+3}, 列{header_name} = '{cell_str}' (パターン: {pattern})", "WARNING")
                            suspicious_count += 1
        
        self.log_debug(f"疑わしいパターン総数: {suspicious_count}")
    
    def run_reproduction_test(self):
        """
        再現テストの実行
        """
        self.log_debug("=== カラムズレ問題再現テスト開始 ===")
        
        # テスト対象ファイル
        test_files = [
            r"G:\共有ドライブ\k_40100_福岡県_北九州市_01\k_北九州市\99_共通\99_資料\02_返礼品関係\02_返礼品シート\企業名\h_株式会社ハマダ\2023-08-03_返礼品登録シート_株式会社ハマダ(赤身スライス900g定期便)（市確認）.xlsx",
            r"G:\共有ドライブ\★OD\99_商品管理\不整合データ\h_株式会社ハマダ\2023-08-03_返礼品登録シート_株式会社ハマダ(赤身スライス900g定期便)（市確認）.xlsx"
        ]
        
        results = {}
        
        for file_path in test_files:
            if os.path.exists(file_path):
                result = self.test_problematic_file(file_path)
                results[os.path.basename(file_path)] = result
            else:
                self.log_debug(f"テストファイルが見つかりません: {file_path}", "WARNING")
        
        # 既存のPhase処理結果との比較
        self.compare_with_existing_results()
        
        return results
    
    def compare_with_existing_results(self):
        """
        既存の処理結果との比較
        """
        self.log_debug("=== 既存結果との比較 ===")
        
        # Phase3の結果をチェック
        phase3_dir = f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase3\\HARV\\{self.municipality}"
        
        if os.path.exists(phase3_dir):
            # all_collect.xlsxをチェック
            all_collect_file = os.path.join(phase3_dir, "all_collect.xlsx")
            if os.path.exists(all_collect_file):
                try:
                    df = pd.read_excel(all_collect_file)
                    
                    # ハマダ関連のレコードを抽出
                    hamada_data = df[df['ファイル名'].astype(str).str.contains('ハマダ', na=False)]
                    self.log_debug(f"既存結果: ハマダ関連レコード {len(hamada_data)}件")
                    
                    # 疑わしいデータをチェック
                    suspicious_in_existing = 0
                    for col in df.columns:
                        if '商品名' in col:
                            name_issues = df[df[col].astype(str).str.contains('河瀨|担当者|氏名', na=False)]
                            if not name_issues.empty:
                                suspicious_in_existing += len(name_issues)
                                self.log_debug(f"既存結果の問題: {col}列に人名パターン {len(name_issues)}件")
                        
                        elif '内容量' in col or '容量' in col:
                            company_issues = df[df[col].astype(str).str.contains('株式会社|有限会社', na=False)]
                            if not company_issues.empty:
                                suspicious_in_existing += len(company_issues)
                                self.log_debug(f"既存結果の問題: {col}列に会社名パターン {len(company_issues)}件")
                    
                    self.log_debug(f"既存結果の問題総数: {suspicious_in_existing}件")
                    
                except Exception as e:
                    self.log_debug(f"既存結果の読み込みエラー: {e}", "ERROR")
    
    def save_debug_log(self):
        """
        デバッグログを保存
        """
        output_file = "debug/column_misalignment_reproduction_log.txt"
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write("\n".join(self.debug_info))
            print(f"デバッグログを保存しました: {output_file}")
        except Exception as e:
            print(f"ログ保存エラー: {e}")

def main():
    reproducer = ColumnMisalignmentReproducer()
    results = reproducer.run_reproduction_test()
    reproducer.save_debug_log()
    return results

if __name__ == "__main__":
    main()
