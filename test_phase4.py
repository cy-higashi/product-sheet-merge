import os
from openpyxl import load_workbook, Workbook

def is_effectively_empty(cell_value, col_index=None):
    """
    セルの値がNoneまたは空文字の場合はTrueを返す。
    また、特定の列についてはエラー・プレースホルダー値も空と見なす。
    
    例:
      - Y列（col_index==25）："発送温度帯を選択してください"
      - AN列（col_index==40）："配送方法を選択してください"
    """
    if cell_value is None or cell_value == '':
        return True
    if col_index == 25 and cell_value == "発送温度帯を選択してください":
        return True
    if col_index == 40 and cell_value == "配送方法を選択してください":
        return True
    return False

def combine_rows(row1, row2, col_count):
    """
    2行分のレコードを統合します。
    すべてのセルについて、両行とも非空の場合は文字列として連結、
    片方のみ非空ならその値を、両方空なら空文字とします。
    なお、両者が完全一致している場合は、重複を防いで1つだけ採用します。
    """
    combined = []
    for i in range(col_count):
        val1 = row1[i] if row1[i] is not None else ""
        val2 = row2[i] if row2[i] is not None else ""
        # 両方空なら結果も空
        if not val1 and not val2:
            combined.append("")
        # 両方非空の場合
        elif val1 and val2:
            # 完全一致なら一方のみ
            if str(val1) == str(val2):
                combined.append(val1)
            else:
                # 両方異なる場合は連結（そのまま連結）
                combined.append(str(val1) + str(val2))
        else:
            # 片方だけ非空ならその値
            combined.append(val1 or val2)
    return combined

def process_file(file_path, output_path):
    # data_only=True でワークブックを読み込み、計算結果（値のみ）を取得
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # ① ワークシート全体を値のマトリックスにコピー（値のみ貼り付け）
    data = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        data.append([cell.value for cell in row])
    
    # ② 結合セルの解除：各結合範囲について、上位セルの値で全セルを埋める
    for merged_range in ws.merged_cells.ranges:
        r1, r2 = merged_range.min_row, merged_range.max_row
        c1, c2 = merged_range.min_col, merged_range.max_col
        top_left_value = data[r1 - 1][c1 - 1]
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                data[r - 1][c - 1] = top_left_value

    # ③ ヘッダー行とデータ行に分割
    # ※ 本来の構造では、ヘッダーが1行目であっても、今回の要求は「B列のセル数＝行数」にするため、
    #     連続する行（キーはB列の値）を統合する処理を全データに適用します。
    header = data[0]
    col_count = len(header)
    data_rows = data[1:]  # ヘッダー以降のすべての行

    # ④ B列（2列目）の値をキーとして、連続する行であれば統合する
    merged_rows = []
    if data_rows:
        current_record = data_rows[0]
        for next_row in data_rows[1:]:
            # 連続する行かどうかはB列（インデックス1）の値で判断
            if next_row[1] == current_record[1]:
                # 既に連結されている行同士は、各列のテキストを連結する
                current_record = combine_rows(current_record, next_row, col_count)
            else:
                merged_rows.append(current_record)
                current_record = next_row
        merged_rows.append(current_record)

    # ⑤ 新しいワークブックに出力
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(header)
    for record in merged_rows:
        new_ws.append(record)
    
    new_wb.save(output_path)
    print(f"正常終了: {os.path.basename(output_path)} を保存しました。")

def main():
    municipality = os.environ.get("MUNICIPALITY_NAME")
    if not municipality:
        print("MUNICIPALITY_NAME 環境変数が設定されていません。")
        return

    base_dir = os.path.join(
        r'G:\共有ドライブ\★OD\99_商品管理\DATA\Phase3\HARV',
        municipality
    )

    if not os.path.exists(base_dir):
        print(f"指定フォルダが存在しません: {base_dir}")
        return

    print(f"MUNICIPALITY_NAME = {municipality}")
    print(f"処理対象フォルダ: {base_dir}")

    files = [f for f in os.listdir(base_dir) if f.startswith("PAT") and f.endswith(".xlsx")]
    if not files:
        print("PATで始まるxlsxファイルが見つかりません。")
        return

    for file in files:
        input_path = os.path.join(base_dir, file)
        output_path = os.path.join(base_dir, file.replace(".xlsx", "_normalized.xlsx"))
        try:
            process_file(input_path, output_path)
        except Exception as e:
            print(f"エラー: {file} の処理に失敗しました → {e}")

if __name__ == "__main__":
    main()
