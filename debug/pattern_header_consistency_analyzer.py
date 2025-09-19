#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
パターン内ヘッダー一致処理の脆弱性分析

特に PAT0004（ハマダ対象ファイル）での列ズレ原因を特定
"""

import os
import pandas as pd
from openpyxl import load_workbook
import json
from datetime import datetime
from collections import defaultdict

class PatternHeaderConsistencyAnalyzer:
    def __init__(self, municipality_name="2025-09-09_福岡県北九州市"):
        self.municipality = municipality_name
        self.target_path = "G:\\共有ドライブ\\k_40100_福岡県_北九州市_01\\k_北九州市\\99_共通\\99_資料\\02_返礼品関係\\02_返礼品シート\\企業名"
        self.base_dirs = {
            'phase1': f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase1\\HARV\\{municipality_name}",
            'phase2': f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase2\\HARV\\{municipality_name}",
            'phase3': f"G:\\共有ドライブ\\★OD\\99_商品管理\\DATA\\Phase3\\HARV\\{municipality_name}"
        }
        self.results = {}
    
    def analyze_pat0004_files(self):
        """PAT0004に分類されたファイルの詳細分析"""
        print("\n=== PAT0004ファイル詳細分析 ===")
        
        # Phase1のファイル別パターンから PAT0004 を抽出
        phase1_file_pattern = os.path.join(self.base_dirs['phase1'], f"{self.municipality}_ファイル別パターン.xlsx")
        
        if not os.path.exists(phase1_file_pattern):
            print(f"Phase1ファイルが存在しません: {phase1_file_pattern}")
            return {}
        
        df = pd.read_excel(phase1_file_pattern)
        pat0004_files = df[df['パターン名'] == 'PAT0004']
        
        print(f"PAT0004に分類されたファイル数: {len(pat0004_files)}")
        
        pat0004_analysis = {}
        
        for _, row in pat0004_files.iterrows():
            folder_name = row['フォルダ名']
            file_name = row['ファイル名']
            file_path = os.path.join(self.target_path, folder_name, file_name)
            
            print(f"\n--- {file_name} ---")
            
            if not os.path.exists(file_path):
                print(f"  ファイルが存在しません: {file_path}")
                continue
            
            # 各ファイルのヘッダー構造を詳細分析
            file_analysis = self.analyze_single_file_headers(file_path)
            pat0004_analysis[file_name] = file_analysis
        
        self.results['pat0004_analysis'] = pat0004_analysis
        return pat0004_analysis
    
    def analyze_single_file_headers(self, file_path):
        """単一ファイルのヘッダー構造を詳細分析"""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # ワークシート全体を2次元リストに変換
            data = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                data.append([cell.value for cell in row])
            
            # 結合セル情報
            merged_ranges = list(ws.merged_cells.ranges)
            
            # 結合セルの解除（値を展開）
            for merged_range in merged_ranges:
                r1, r2 = merged_range.min_row, merged_range.max_row
                c1, c2 = merged_range.min_col, merged_range.max_col
                top_left_value = data[r1 - 1][c1 - 1]
                for r in range(r1, r2 + 1):
                    for c in range(c1, c2 + 1):
                        data[r - 1][c - 1] = top_left_value
            
            # ヘッダー基準位置の検出（merge.pyと同じロジック）
            header_row_index, header_col_start = self.find_header_base_position_detailed(data)
            
            if header_row_index is None or header_col_start is None:
                return {
                    'error': 'ヘッダー基準位置が見つからない',
                    'merged_ranges_count': len(merged_ranges)
                }
            
            # ヘッダー構造の抽出
            main_row = data[header_row_index] if header_row_index < len(data) else []
            sub_row = data[header_row_index + 1] if header_row_index + 1 < len(data) else []
            
            # 各列のヘッダー情報を詳細収集
            header_details = []
            max_col = max(len(main_row), len(sub_row))
            
            for col_idx in range(header_col_start, min(max_col, header_col_start + 50)):  # 最大50列
                main_header = main_row[col_idx] if col_idx < len(main_row) else None
                sub_header = sub_row[col_idx] if col_idx < len(sub_row) else None
                
                # ヘッダー名生成（merge.pyと同じロジック）
                header_name = self.create_hierarchical_header(main_header, sub_header, col_idx, header_col_start)
                
                # サンプルデータ（最初の5行）
                sample_data = []
                for data_row_idx in range(header_row_index + 2, min(header_row_index + 7, len(data))):
                    if data_row_idx < len(data) and col_idx < len(data[data_row_idx]):
                        sample_data.append(data[data_row_idx][col_idx])
                    else:
                        sample_data.append(None)
                
                header_details.append({
                    'column_index': col_idx,
                    'main_header': main_header,
                    'sub_header': sub_header,
                    'generated_header_name': header_name,
                    'sample_data': sample_data
                })
            
            return {
                'header_row_index': header_row_index,
                'header_col_start': header_col_start,
                'merged_ranges_count': len(merged_ranges),
                'header_details': header_details,
                'main_row_length': len(main_row),
                'sub_row_length': len(sub_row)
            }
            
        except Exception as e:
            return {'error': str(e)}
    
    def find_header_base_position_detailed(self, data):
        """ヘッダー基準位置検出（詳細ログ付き）"""
        header_row_index = None
        header_col_index = None
        
        # 「項目」検索
        item_candidates = []
        for i, row in enumerate(data):
            if len(row) >= 2 and row[1] == "項目":
                item_candidates.append(i)
        
        if item_candidates:
            header_row_index = item_candidates[0]
        else:
            return None, None
        
        # 「返礼品コード」検索（下方向探索）
        search_window_rows = [header_row_index]
        for k in range(1, 6):
            if header_row_index + k < len(data):
                search_window_rows.append(header_row_index + k)
        
        for r in search_window_rows:
            if r < len(data):
                row = data[r]
                for j, cell in enumerate(row):
                    if cell and "返礼品コード" in str(cell):
                        header_col_index = j
                        break
            if header_col_index is not None:
                break
        
        return header_row_index, header_col_index
    
    def create_hierarchical_header(self, main_header, sub_header, col_index, header_col_start):
        """階層ヘッダー名生成（merge.pyと同じロジック）"""
        def normalize_header_text(text):
            if not text:
                return ""
            text = str(text).strip()
            import re
            text = re.sub(r'\s+', '', text)
            return text
        
        main_clean = normalize_header_text(main_header) if main_header else ""
        sub_clean = normalize_header_text(sub_header) if sub_header else ""
        
        if main_clean and sub_clean and sub_clean not in ["必須", "任意"]:
            return f"{main_clean}:{sub_clean}"
        elif main_clean:
            return main_clean
        elif sub_clean and sub_clean not in ["必須", "任意"]:
            return sub_clean
        else:
            col_relative = col_index - header_col_start + 1
            return f"空列{col_relative}"
    
    def compare_header_consistency(self):
        """PAT0004内でのヘッダー一致性比較"""
        print("\n=== PAT0004内ヘッダー一致性比較 ===")
        
        if 'pat0004_analysis' not in self.results:
            print("PAT0004分析結果がありません")
            return
        
        pat0004_files = self.results['pat0004_analysis']
        
        # 各ファイルのヘッダー名リストを収集
        header_sets = {}
        for file_name, analysis in pat0004_files.items():
            if 'error' in analysis:
                print(f"{file_name}: エラー - {analysis['error']}")
                continue
            
            headers = [detail['generated_header_name'] for detail in analysis.get('header_details', [])]
            header_sets[file_name] = headers
        
        # ヘッダー名の一致度分析
        if len(header_sets) < 2:
            print("比較対象ファイルが不足")
            return
        
        # 全ファイル共通のヘッダー
        all_headers = set()
        for headers in header_sets.values():
            all_headers.update(headers)
        
        common_headers = set(all_headers)
        for headers in header_sets.values():
            common_headers &= set(headers)
        
        print(f"全ヘッダー数: {len(all_headers)}")
        print(f"共通ヘッダー数: {len(common_headers)}")
        print(f"一致率: {len(common_headers)/len(all_headers)*100:.1f}%")
        
        # ファイル間差異の詳細
        inconsistencies = []
        file_names = list(header_sets.keys())
        
        for i in range(len(file_names)):
            for j in range(i+1, len(file_names)):
                file1, file2 = file_names[i], file_names[j]
                headers1, headers2 = set(header_sets[file1]), set(header_sets[file2])
                
                only_in_1 = headers1 - headers2
                only_in_2 = headers2 - headers1
                
                if only_in_1 or only_in_2:
                    inconsistencies.append({
                        'file1': file1,
                        'file2': file2,
                        'only_in_file1': list(only_in_1),
                        'only_in_file2': list(only_in_2)
                    })
        
        self.results['header_consistency'] = {
            'total_headers': len(all_headers),
            'common_headers': len(common_headers),
            'consistency_rate': len(common_headers)/len(all_headers)*100,
            'inconsistencies': inconsistencies
        }
        
        # 不一致の詳細表示
        print("\n不一致詳細:")
        for inc in inconsistencies[:3]:  # 最初の3件
            print(f"  {inc['file1']} vs {inc['file2']}:")
            if inc['only_in_file1']:
                print(f"    {inc['file1']}のみ: {inc['only_in_file1'][:5]}")  # 最初の5個
            if inc['only_in_file2']:
                print(f"    {inc['file2']}のみ: {inc['only_in_file2'][:5]}")  # 最初の5個
    
    def analyze_problematic_columns(self):
        """問題のある列（商品名/内容量）の具体分析"""
        print("\n=== 問題列の具体分析 ===")
        
        if 'pat0004_analysis' not in self.results:
            return
        
        problematic_patterns = []
        
        for file_name, analysis in self.results['pat0004_analysis'].items():
            if 'error' in analysis:
                continue
            
            for detail in analysis.get('header_details', []):
                header_name = detail['generated_header_name']
                sample_data = detail['sample_data']
                
                # 商品名列に人名系データが混入していないかチェック
                if '商品名' in header_name:
                    for sample in sample_data:
                        if sample and any(keyword in str(sample) for keyword in ['担当者', 'ご担当者', '様', 'さん']):
                            problematic_patterns.append({
                                'file': file_name,
                                'column': header_name,
                                'issue': '商品名列に人名データ',
                                'sample_value': sample
                            })
                
                # 内容量列に会社名系データが混入していないかチェック
                if '内容量' in header_name or '容量' in header_name:
                    for sample in sample_data:
                        if sample and any(keyword in str(sample) for keyword in ['株式会社', '有限会社', '合同会社']):
                            problematic_patterns.append({
                                'file': file_name,
                                'column': header_name,
                                'issue': '内容量列に会社名データ',
                                'sample_value': sample
                            })
        
        self.results['problematic_patterns'] = problematic_patterns
        
        print(f"問題パターン検出数: {len(problematic_patterns)}")
        for pattern in problematic_patterns[:5]:  # 最初の5件
            print(f"  {pattern['file']}: {pattern['column']} - {pattern['issue']}")
            print(f"    サンプル値: {pattern['sample_value']}")
    
    def run_full_analysis(self):
        """完全分析の実行"""
        print("=== PAT0004ヘッダー一致性分析開始 ===")
        print(f"分析開始時刻: {datetime.now()}")
        
        # PAT0004ファイルの詳細分析
        self.analyze_pat0004_files()
        
        # ヘッダー一致性比較
        self.compare_header_consistency()
        
        # 問題列の具体分析
        self.analyze_problematic_columns()
        
        # 結果保存
        output_file = "debug/pattern_header_consistency_analysis.json"
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(self.results, f, ensure_ascii=False, indent=2, default=str)
            print(f"\n分析結果を保存しました: {output_file}")
        except Exception as e:
            print(f"結果保存エラー: {e}")
        
        print("\n=== 分析完了 ===")
        return self.results

def main():
    analyzer = PatternHeaderConsistencyAnalyzer()
    results = analyzer.run_full_analysis()
    return results

if __name__ == "__main__":
    main()
